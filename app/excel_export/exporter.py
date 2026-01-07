# -*- coding: utf-8 -*-
"""
Exportador de relatórios de evolução para Excel
"""

from io import BytesIO
from typing import Dict, Any, List, Optional
from datetime import datetime
import logging
import tempfile
import os

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

from app.excel_export.data_transformer import DataTransformer
from app.excel_export.formatters import ExcelFormatter
from app.excel_export.matplotlib_chart_builder import MatplotlibChartBuilder
from app.services.evaluation_comparison_service import EvaluationComparisonService
from app.models.test import Test
from app.models.city import City

logger = logging.getLogger(__name__)


class ExcelEvolutionExporter:
    """Classe principal para exportação de relatórios de evolução para Excel"""
    
    def __init__(self):
        self.workbook = None
        self.transformed_data = None
        self.temp_files = []  # Lista para rastrear arquivos temporários
    
    def export(self, test_ids: List[str], municipality: Optional[str] = None, 
               state: Optional[str] = None, department: Optional[str] = None) -> BytesIO:
        """
        Exporta relatório de evolução para Excel
        
        Args:
            test_ids: Lista de IDs das avaliações
            municipality: Nome do município (opcional)
            state: Nome do estado (opcional)
            department: Nome do departamento/secretaria (opcional)
            
        Returns:
            BytesIO com arquivo Excel
        """
        try:
            # Buscar dados de comparação
            comparison_data = EvaluationComparisonService.compare_evaluations(test_ids)
            if not comparison_data:
                raise ValueError("Não foi possível obter dados de comparação")
            
            # Transformar dados para formato tabular
            self.transformed_data = DataTransformer.transform_comparison_data(comparison_data)
            if not self.transformed_data:
                raise ValueError("Não foi possível transformar dados")
            
            # Criar workbook
            self.workbook = Workbook()
            self.workbook.remove(self.workbook.active)  # Remover sheet padrão
            
            # Buscar informações adicionais das avaliações
            tests = Test.query.filter(Test.id.in_(test_ids)).all()
            evaluations_info = comparison_data.get('evaluations', [])
            
            # Tentar obter município e estado das avaliações se não fornecidos
            if not municipality or not state:
                municipality, state = self._extract_location_info(tests)
            
            if not department:
                department = "SECRETARIA MUNICIPAL DE EDUCAÇÃO"
            
            # Criar abas
            self._create_cover_sheet(municipality, state, department, evaluations_info)
            self._create_general_sheet()
            
            # Criar aba de participação
            self._create_participation_sheet()
            
            # Criar abas por disciplina
            subjects = self.transformed_data.get('subjects', {})
            for subject_name in subjects.keys():
                self._create_subject_sheet(subject_name)
            
            # Criar aba de níveis
            self._create_classification_sheet()
            
            # Salvar em BytesIO
            output = BytesIO()
            self.workbook.save(output)
            output.seek(0)
            
            # Limpar arquivos temporários
            self._cleanup_temp_files()
            
            return output
            
        except Exception as e:
            logger.error(f"Erro ao exportar relatório Excel: {str(e)}", exc_info=True)
            # Limpar arquivos temporários mesmo em caso de erro
            self._cleanup_temp_files()
            raise
    
    def _cleanup_temp_files(self):
        """Remove arquivos temporários criados durante a exportação"""
        for temp_file in self.temp_files:
            try:
                if os.path.exists(temp_file):
                    os.remove(temp_file)
                    logger.debug(f"Arquivo temporário removido: {temp_file}")
            except Exception as e:
                logger.warning(f"Erro ao remover arquivo temporário {temp_file}: {str(e)}")
        self.temp_files = []
    
    def _extract_location_info(self, tests: List[Test]) -> tuple:
        """Extrai informações de localização das avaliações"""
        municipality = None
        state = None
        
        for test in tests:
            if test.municipalities and isinstance(test.municipalities, list) and len(test.municipalities) > 0:
                # Tentar buscar cidade
                city_id = test.municipalities[0]
                city = City.query.get(city_id)
                if city:
                    municipality = city.name
                    state = city.state
                    break
        
        return municipality or "SÃO MIGUEL DOS CAMPOS", state or "ALAGOAS"
    
    def _create_cover_sheet(self, municipality: str, state: str, department: str,
                          evaluations_info: List[Dict]):
        """Cria aba de capa"""
        sheet = self.workbook.create_sheet("Capa")
        
        # Logo no topo
        try:
            import os
            logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'assets', 'afirme_logo.png')
            if os.path.exists(logo_path):
                img = Image(logo_path)
                img.width = 150
                img.height = 150
                sheet.add_image(img, 'A1')
        except Exception as e:
            logger.warning(f"Erro ao adicionar logo na capa: {e}")
        
        # Título principal (abaixar para dar espaço ao logo)
        ExcelFormatter.merge_and_center(
            sheet, "A6:H6", 
            "RELATÓRIO DE EVOLUÇÃO DAS AVALIAÇÕES",
            font=ExcelFormatter.TITLE_FONT
        )
        
        # Informações do município (ajustar row para dar espaço ao logo)
        row = 8
        sheet.cell(row, 1).value = f"{municipality} - {state}"
        ExcelFormatter.format_subtitle_cell(sheet.cell(row, 1))
        sheet.merge_cells(f"A{row}:H{row}")
        
        row += 1
        sheet.cell(row, 1).value = department
        ExcelFormatter.format_subtitle_cell(sheet.cell(row, 1))
        sheet.merge_cells(f"A{row}:H{row}")
        
        # Lista de avaliações
        row += 3
        sheet.cell(row, 1).value = "AVALIAÇÕES COMPARADAS:"
        ExcelFormatter.format_subtitle_cell(sheet.cell(row, 1))
        
        row += 1
        for i, eval_info in enumerate(evaluations_info, 1):
            eval_title = eval_info.get('title', f'Avaliação {i}')
            eval_date = eval_info.get('application_date', '')
            if eval_date:
                try:
                    date_obj = datetime.fromisoformat(eval_date.replace('Z', '+00:00'))
                    eval_date = date_obj.strftime('%d/%m/%Y')
                except:
                    pass
            
            sheet.cell(row, 1).value = f"{i}. {eval_title}"
            if eval_date:
                sheet.cell(row, 2).value = f"({eval_date})"
            ExcelFormatter.format_data_cell(sheet.cell(row, 1))
            ExcelFormatter.format_data_cell(sheet.cell(row, 2))
            row += 1
        
        # Data de geração
        row += 2
        sheet.cell(row, 1).value = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ExcelFormatter.format_data_cell(sheet.cell(row, 1))
        
        # Badge AfirmePlay
        row += 2
        ExcelFormatter.merge_and_center(
            sheet, f"A{row}:H{row}",
            "AFIRMEPLAY SISTEMA DE ENSINO E AVALIAÇÃO",
            font=ExcelFormatter.BOLD_FONT,
            fill=ExcelFormatter.HEADER_FILL
        )
        
        # Ajustar larguras
        ExcelFormatter.auto_adjust_column_width(sheet)
    
    def _create_general_sheet(self):
        """Cria aba com dados gerais - com logo, título e subtítulo"""
        sheet = self.workbook.create_sheet("Geral")
        
        general_data = self.transformed_data.get('general', {})
        evaluations = self.transformed_data.get('evaluations', [])
        
        if not general_data or not evaluations:
            return
        
        # Garantir sempre 3 avaliações
        eval_list = evaluations[:3] if len(evaluations) >= 3 else evaluations
        while len(eval_list) < 3:
            eval_list.append({'title': f'Avaliação {len(eval_list) + 1}', 'id': f'fake_{len(eval_list) + 1}'})
        
        row = 1
        
        # Logo no topo
        try:
            import os
            logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'assets', 'afirme_logo.png')
            if os.path.exists(logo_path):
                img = Image(logo_path)
                img.width = 150
                img.height = 150
                sheet.add_image(img, 'A1')
                row = 5  # Pular espaço para logo
        except Exception as e:
            logger.warning(f"Erro ao adicionar logo: {e}")
            row = 2
        
        # Título "Afirme Play"
        ExcelFormatter.merge_and_center(
            sheet, f"A{row}:H{row}",
            "AFIRME PLAY",
            font=ExcelFormatter.TITLE_FONT
        )
        
        # Subtítulo
        row += 1
        ExcelFormatter.merge_and_center(
            sheet, f"A{row}:H{row}",
            "Relatório de Desempenho Educacional",
            font=ExcelFormatter.SUBTITLE_FONT
        )
        
        # Espaço
        row += 3
        
        # BLOCO 1 - Nota Geral
        row = self._create_metric_table(
            sheet, row_start=row,
            title="NOTA GERAL",
            values=general_data.get('average_grades', [])[:3],  # Garantir 3 valores
            variations=general_data.get('grade_variations', [])[:2],  # Garantir 2 variações
            directions=general_data.get('grade_directions', [])[:2],
            evaluations=eval_list,
            y_axis_max=10,
            y_axis_min=0
        )
        
        # BLOCO 2 - Proficiência Geral
        # row já contém o espaço grande depois do gráfico anterior
        row = self._create_metric_table(
            sheet, row_start=row,
            title="PROFICIÊNCIA GERAL",
            values=general_data.get('average_proficiencies', [])[:3],
            variations=general_data.get('proficiency_variations', [])[:2],
            directions=general_data.get('proficiency_directions', [])[:2],
            evaluations=eval_list,
            y_axis_max=425,
            y_axis_min=0
        )
        
        ExcelFormatter.auto_adjust_column_width(sheet)
    
    def _create_subject_sheet(self, subject_name: str):
        """Cria aba para uma disciplina - sempre 3 avaliações"""
        # Limitar nome da aba (Excel tem limite de 31 caracteres)
        sheet_name = subject_name[:31] if len(subject_name) > 31 else subject_name
        sheet = self.workbook.create_sheet(sheet_name)
        
        subject_data = self.transformed_data.get('subjects', {}).get(subject_name, {})
        evaluations = self.transformed_data.get('evaluations', [])
        
        if not subject_data or not evaluations:
            return
        
        # Garantir sempre 3 avaliações
        eval_list = evaluations[:3] if len(evaluations) >= 3 else evaluations
        while len(eval_list) < 3:
            eval_list.append({'title': f'Avaliação {len(eval_list) + 1}', 'id': f'fake_{len(eval_list) + 1}'})
        
        # Tabela de Nota por Disciplina
        row = self._create_metric_table(
            sheet, row_start=1,
            title=f"NOTA - {subject_name.upper()}",
            values=subject_data.get('average_grades', [])[:3],
            variations=subject_data.get('grade_variations', [])[:2],
            directions=subject_data.get('grade_directions', [])[:2],
            evaluations=eval_list,
            y_axis_max=10,
            y_axis_min=0
        )
        
        # Tabela de Proficiência por Disciplina
        # row já contém o espaço grande depois do gráfico anterior
        row = self._create_metric_table(
            sheet, row_start=row,
            title=f"PROFICIÊNCIA - {subject_name.upper()}",
            values=subject_data.get('average_proficiencies', [])[:3],
            variations=subject_data.get('proficiency_variations', [])[:2],
            directions=subject_data.get('proficiency_directions', [])[:2],
            evaluations=eval_list,
            y_axis_max=425,
            y_axis_min=0
        )
        
        ExcelFormatter.auto_adjust_column_width(sheet)
    
    def _create_participation_sheet(self):
        """Cria aba com dados de participação geral e por escola"""
        sheet = self.workbook.create_sheet("Participação")
        
        participation_data = self.transformed_data.get('participation', {})
        evaluations = self.transformed_data.get('evaluations', [])
        
        if not participation_data or not evaluations:
            return
        
        # Garantir sempre 3 avaliações
        eval_list = evaluations[:3] if len(evaluations) >= 3 else evaluations
        while len(eval_list) < 3:
            eval_list.append({'title': f'Avaliação {len(eval_list) + 1}', 'id': f'fake_{len(eval_list) + 1}'})
        
        row = 1
        
        # Logo no topo
        try:
            import os
            logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'assets', 'afirme_logo.png')
            if os.path.exists(logo_path):
                img = Image(logo_path)
                img.width = 150
                img.height = 150
                sheet.add_image(img, 'A1')
                row = 5
        except Exception as e:
            logger.warning(f"Erro ao adicionar logo: {e}")
            row = 2
        
        # Título
        ExcelFormatter.merge_and_center(
            sheet, f"A{row}:H{row}",
            "TAXA DE PARTICIPAÇÃO",
            font=ExcelFormatter.TITLE_FONT
        )
        
        row += 2
        
        # Seção 1: Taxa Geral de Participação
        general_data = participation_data.get('general', {})
        if general_data:
            row = self._create_general_participation_table(
                sheet, row_start=row,
                participation_rates=general_data.get('participation_rates', [])[:3],
                total_students=general_data.get('total_students', [])[:3],
                participating_students=general_data.get('participating_students', [])[:3],
                variations=general_data.get('variations', [])[:2],
                directions=general_data.get('directions', [])[:2],
                evaluations=eval_list
            )
        
        # Seção 2: Taxa por Escola (uma tabela por avaliação)
        by_school_data = participation_data.get('by_school', {})
        if by_school_data:
            row += 5  # Espaço entre seções
            
            for i, eval_info in enumerate(eval_list):
                eval_key = f'evaluation_{i+1}'
                schools_data = by_school_data.get(eval_key, {})
                
                if schools_data:
                    row = self._create_school_participation_table(
                        sheet,
                        row_start=row,
                        evaluation_info=eval_info,
                        schools_data=schools_data
                    )
                    row += 5  # Espaço entre avaliações
        
        ExcelFormatter.auto_adjust_column_width(sheet)
    
    def _create_general_participation_table(self, sheet, row_start: int, participation_rates: List[float],
                                           total_students: List[int], participating_students: List[int],
                                           variations: List[float], directions: List[str],
                                           evaluations: List[Dict]) -> int:
        """Cria tabela de participação geral com gráfico"""
        col = 1
        row = row_start
        
        # Garantir sempre 3 avaliações
        eval_list = evaluations[:3] if len(evaluations) >= 3 else evaluations
        while len(eval_list) < 3:
            eval_list.append({'title': f'Avaliação {len(eval_list) + 1}', 'id': f'fake_{len(eval_list) + 1}'})
        
        # Título
        sheet.cell(row, col).value = "TAXA GERAL DE PARTICIPAÇÃO"
        ExcelFormatter.format_subtitle_cell(sheet.cell(row, col))
        sheet.merge_cells(f"A{row}:{get_column_letter(1 + 3 + 2 + 1)}{row}")
        
        # Cabeçalho
        row += 2
        sheet.cell(row, col).value = "Avaliação"
        ExcelFormatter.format_header_cell(sheet.cell(row, col))
        
        # Colunas de valores (sempre 3)
        data_start_col = col + 1
        for i in range(3):
            col += 1
            eval_title = eval_list[i].get('title', f'Avaliação {i+1}')[:15] if i < len(eval_list) else f'Avaliação {i+1}'
            sheet.cell(row, col).value = eval_title
            ExcelFormatter.format_header_cell(sheet.cell(row, col))
        
        data_end_col = col
        
        # Colunas de variações (sempre 2)
        var_start_col = col + 1
        for i in range(2):
            col += 1
            sheet.cell(row, col).value = f"Var {i+1}→{i+2} (%)"
            ExcelFormatter.format_header_cell(sheet.cell(row, col))
        
        col += 1
        sheet.cell(row, col).value = "Direção"
        ExcelFormatter.format_header_cell(sheet.cell(row, col))
        
        # Linha de Taxa (%)
        row += 1
        col = 1
        sheet.cell(row, col).value = "Taxa (%)"
        ExcelFormatter.format_data_cell(sheet.cell(row, col))
        
        # Contar quantas colunas realmente têm dados
        actual_data_cols = 0
        for i, rate in enumerate(participation_rates):
            col += 1
            if rate is None:
                ExcelFormatter.format_percentage_cell(sheet.cell(row, col), 0, 'stable')
                if i < len(participation_rates) - 1:
                    actual_data_cols = i + 1
            else:
                # Taxa de participação vem como porcentagem (ex: 95.0)
                # format_percentage_cell com formato '0.00"%"' apenas adiciona %, não multiplica
                # Então precisamos dividir por 100 para mostrar corretamente
                ExcelFormatter.format_percentage_cell(sheet.cell(row, col), rate / 100.0, 'stable')
                actual_data_cols = i + 1
        
        if actual_data_cols < 2:
            actual_data_cols = len(participation_rates)
        
        # Linha de variações
        var_row = row + 1
        var_col = var_start_col
        for i, (var_value, direction) in enumerate(zip(variations, directions)):
            ExcelFormatter.format_percentage_cell(sheet.cell(var_row, var_col), var_value, direction)
            var_col += 1
        
        # Direção (primeira variação)
        if directions:
            ExcelFormatter.format_direction_cell(sheet.cell(var_row, var_start_col + len(variations)), directions[0])
        
        # Gráfico usando Matplotlib
        chart_row = var_row + 5
        
        if actual_data_cols >= 2:
            eval_labels = [eval_list[i].get('title', f'Avaliação {i+1}')[:30] if i < len(eval_list) else f'Avaliação {i+1}' 
                          for i in range(actual_data_cols)]
            
            # Criar arquivo temporário para a imagem
            temp_fd, temp_path = tempfile.mkstemp(suffix='.png', prefix='chart_participacao_geral_')
            os.close(temp_fd)
            self.temp_files.append(temp_path)
            
            valid_variations = [v for v in variations if v is not None]
            valid_directions = [d for i, d in enumerate(directions) if i < len(valid_variations)]
            
            image_path = None
            if actual_data_cols >= 2 and valid_variations:
                image_path = MatplotlibChartBuilder.create_composed_chart_image(
                    labels=eval_labels,
                    values=participation_rates[:actual_data_cols],
                    variations=valid_variations,
                    directions=valid_directions,
                    chart_title="TAXA GERAL DE PARTICIPAÇÃO",
                    y_axis_title="Taxa (%)",
                    y_axis_min=0,
                    y_axis_max=100,
                    output_path=temp_path
                )
            else:
                image_path = MatplotlibChartBuilder.create_bar_chart_image(
                    labels=eval_labels,
                    values=participation_rates[:actual_data_cols],
                    chart_title="TAXA GERAL DE PARTICIPAÇÃO",
                    y_axis_title="Taxa (%)",
                    y_axis_min=0,
                    y_axis_max=100,
                    output_path=temp_path
                )
            
            if image_path and os.path.exists(image_path):
                try:
                    img = Image(image_path)
                    img.width = 600
                    img.height = 300
                    img.anchor = f"A{chart_row}"
                    sheet.add_image(img)
                    estimated_chart_end_row = chart_row + 15
                    return estimated_chart_end_row + 10
                except Exception as e:
                    logger.error(f"Erro ao adicionar imagem do gráfico: {str(e)}")
        
        return var_row + 5
    
    def _create_school_participation_table(self, sheet, row_start: int, evaluation_info: Dict,
                                          schools_data: Dict[str, Dict[str, Any]]) -> int:
        """Cria tabela de participação por escola para uma avaliação"""
        col = 1
        row = row_start
        
        eval_title = evaluation_info.get('title', 'Avaliação')
        
        # Título
        sheet.cell(row, col).value = f"TAXA DE PARTICIPAÇÃO POR ESCOLA - {eval_title.upper()}"
        ExcelFormatter.format_subtitle_cell(sheet.cell(row, col))
        sheet.merge_cells(f"A{row}:E{row}")
        
        # Cabeçalho
        row += 2
        sheet.cell(row, 1).value = "Escola"
        ExcelFormatter.format_header_cell(sheet.cell(row, 1))
        sheet.cell(row, 2).value = "Total Alunos"
        ExcelFormatter.format_header_cell(sheet.cell(row, 2))
        sheet.cell(row, 3).value = "Participantes"
        ExcelFormatter.format_header_cell(sheet.cell(row, 3))
        sheet.cell(row, 4).value = "Taxa (%)"
        ExcelFormatter.format_header_cell(sheet.cell(row, 4))
        
        # Ordenar escolas por nome
        sorted_schools = sorted(schools_data.items(), key=lambda x: x[1].get('school_name', ''))
        
        # Dados das escolas
        row += 1
        for school_id, school_info in sorted_schools:
            sheet.cell(row, 1).value = school_info.get('school_name', f'Escola {school_id}')
            ExcelFormatter.format_data_cell(sheet.cell(row, 1))
            
            total = school_info.get('total_students', 0)
            participating = school_info.get('participating_students', 0)
            rate = school_info.get('participation_rate', 0.0)
            
            ExcelFormatter.format_number_cell(sheet.cell(row, 2), total, decimals=0)
            ExcelFormatter.format_number_cell(sheet.cell(row, 3), participating, decimals=0)
            # Taxa de participação vem como porcentagem (ex: 95.0), dividir por 100 para formato Excel
            ExcelFormatter.format_percentage_cell(sheet.cell(row, 4), rate / 100.0, 'stable')
            
            row += 1
        
        return row
    
    def _create_classification_sheet(self):
        """Cria aba de níveis de classificação com Composed Chart"""
        sheet = self.workbook.create_sheet("Níveis")
        
        classification_data = self.transformed_data.get('classification', {})
        evaluations = self.transformed_data.get('evaluations', [])
        
        if not classification_data or not evaluations:
            return
        
        # Garantir sempre 3 avaliações
        eval_list = evaluations[:3] if len(evaluations) >= 3 else evaluations
        while len(eval_list) < 3:
            eval_list.append({'title': f'Avaliação {len(eval_list) + 1}', 'id': f'fake_{len(eval_list) + 1}'})
        
        class_data = classification_data.get('classification_data', {})
        variations = classification_data.get('variations', {})
        directions = classification_data.get('directions', {})
        
        levels = ['Abaixo do Básico', 'Básico', 'Adequado', 'Avançado']
        
        # Título
        sheet.cell(1, 1).value = "DISTRIBUIÇÃO DE ALUNOS POR NÍVEL"
        ExcelFormatter.format_subtitle_cell(sheet.cell(1, 1))
        sheet.merge_cells(f"A1:{get_column_letter(1 + 3 + 2)}1")
        
        # Cabeçalho
        row = 3
        col = 1
        sheet.cell(row, col).value = "Nível"
        ExcelFormatter.format_header_cell(sheet.cell(row, col))
        
        # Colunas de avaliações (sempre 3)
        data_start_col = col + 1
        for i in range(3):
            col += 1
            eval_title = eval_list[i].get('title', f'Avaliação {i+1}')[:20] if i < len(eval_list) else f'Avaliação {i+1}'
            sheet.cell(row, col).value = eval_title
            ExcelFormatter.format_header_cell(sheet.cell(row, col))
        
        data_end_col = col
        
        # Colunas de variações (sempre 2)
        var_start_col = col + 1
        for i in range(2):
            col += 1
            sheet.cell(row, col).value = f"Var {i+1}→{i+2} (%)"
            ExcelFormatter.format_header_cell(sheet.cell(row, col))
        
        # Dados
        row += 1
        data_row_start = row
        max_value = 0
        
        for level in levels:
            col = 1
            sheet.cell(row, col).value = level
            ExcelFormatter.format_data_cell(sheet.cell(row, col))
            
            # Valores por avaliação (sempre 3)
            values = class_data.get(level, [])[:3]
            while len(values) < 3:
                values.append(0)
            
            for i, value in enumerate(values):
                col += 1
                ExcelFormatter.format_number_cell(sheet.cell(row, col), value, decimals=0)
                if value and value > max_value:
                    max_value = value
            
            # Variações (sempre 2)
            var_values = variations.get(level, [])[:2]
            var_directions = directions.get(level, [])[:2]
            while len(var_values) < 2:
                var_values.append(0)
                var_directions.append('stable')
            
            for i, (var_value, direction) in enumerate(zip(var_values, var_directions)):
                col += 1
                ExcelFormatter.format_percentage_cell(sheet.cell(row, col), var_value, direction)
            
            row += 1
        
        data_row_end = row - 1
        
        # Calcular eixo Y máximo dinâmico
        y_axis_max = max_value + 5 if max_value > 0 else 10
        y_axis_min = 0
        
        # Criar Composed Chart para cada nível
        # O gráfico será posicionado após a linha de variações
        # Primeiro vamos criar a linha de totais e variações, depois calcular chart_row
        
        # Para cada nível, criar um gráfico combinado
        # Vamos criar um gráfico agregado mostrando todos os níveis
        # Criar range de barras (soma de todos os níveis por avaliação)
        # Criar range de linha (variação média)
        
        # Range de categorias (nomes das avaliações no cabeçalho)
        header_row = 3  # Linha do cabeçalho (linha 3 após título)
        categories_range = f"{get_column_letter(data_start_col)}{header_row}:{get_column_letter(data_end_col)}{header_row}"
        
        # Para simplificar, vamos criar um gráfico por nível ou um gráfico agregado
        # Criando gráfico agregado: soma de todos os níveis
        # Primeiro, criar linha de totais
        total_row = row
        col = 1
        sheet.cell(total_row, col).value = "TOTAL"
        ExcelFormatter.format_data_cell(sheet.cell(total_row, col))
        
        # Calcular totais por avaliação (sempre 3)
        totals = [0, 0, 0]
        actual_totals_count = 0
        for level in levels:
            values = class_data.get(level, [])[:3]
            for i, val in enumerate(values):
                if i < 3:
                    if val is not None:
                        totals[i] = totals[i] + (val or 0)
                        if i >= actual_totals_count:
                            actual_totals_count = i + 1
        
        bar_data_start_col = data_start_col
        for i, total in enumerate(totals):
            col += 1
            ExcelFormatter.format_number_cell(sheet.cell(total_row, col), total, decimals=0)
        
        # Ajustar para usar apenas colunas com dados
        if actual_totals_count > 0:
            bar_data_end_col = data_start_col + actual_totals_count - 1
        else:
            bar_data_end_col = col
        
        # Calcular variações dos totais
        total_variations = []
        total_directions = []
        if len(totals) >= 2:
            for i in range(len(totals) - 1):
                if totals[i] > 0:
                    var = ((totals[i+1] - totals[i]) / totals[i]) * 100
                elif totals[i+1] > 0:
                    var = 100.0
                else:
                    var = 0.0
                total_variations.append(var)
                total_directions.append('increase' if var > 0 else ('decrease' if var < 0 else 'stable'))
        
        # Range de barras (totais) - usar apenas colunas com dados
        bar_data_range = f"{get_column_letter(bar_data_start_col)}{total_row}:{get_column_letter(bar_data_end_col)}{total_row}"
        
        # Range de linha - A linha deve conectar os topos das barras usando os mesmos valores
        var_row = total_row + 1
        var_col = var_start_col
        valid_variations = []
        valid_directions = []
        for i, (var_value, direction) in enumerate(zip(total_variations, total_directions)):
            if var_value is not None:
                ExcelFormatter.format_percentage_cell(sheet.cell(var_row, var_col), var_value, direction)
                valid_variations.append(var_value)
                valid_directions.append(direction)
                var_col += 1
        
        # Calcular posição do gráfico: 5 linhas após a linha de variações
        chart_row = var_row + 5
        
        # A linha deve usar os valores das barras (mesmo range) para conectar os topos
        # Verificar quantas colunas de dados temos
        actual_data_cols = bar_data_end_col - bar_data_start_col + 1
        if actual_data_cols >= 2:
            # A linha usa os mesmos valores das barras
            line_data_range = bar_data_range  # Usar os mesmos valores das barras
        else:
            line_data_range = None
        
        # Criar gráfico usando Matplotlib (imagem PNG)
        # Preparar labels das avaliações
        eval_list = self.transformed_data.get('evaluations', [])[:3]
        while len(eval_list) < 3:
            eval_list.append({'title': f'Avaliação {len(eval_list) + 1}', 'id': f'fake_{len(eval_list) + 1}'})
        
        eval_labels = [eval_list[i].get('title', f'Avaliação {i+1}')[:30] if i < len(eval_list) else f'Avaliação {i+1}' 
                      for i in range(actual_data_cols)]
        
        # Criar arquivo temporário para a imagem
        temp_fd, temp_path = tempfile.mkstemp(suffix='.png', prefix='chart_niveis_')
        os.close(temp_fd)
        self.temp_files.append(temp_path)
        
        image_path = None
        if actual_data_cols >= 2 and valid_variations:
            # Criar gráfico combinado com linha conectando os topos das barras
            image_path = MatplotlibChartBuilder.create_composed_chart_image(
                labels=eval_labels,
                values=totals[:actual_data_cols],
                variations=valid_variations,
                directions=valid_directions,
                chart_title="Distribuição de Classificação por Avaliação",
                y_axis_title="Quantidade de Alunos",
                y_axis_min=y_axis_min,
                y_axis_max=y_axis_max,
                output_path=temp_path
            )
        else:
            # Se não houver dados suficientes, criar apenas gráfico de barras
            image_path = MatplotlibChartBuilder.create_bar_chart_image(
                labels=eval_labels,
                values=totals[:actual_data_cols],
                chart_title="Distribuição de Classificação por Avaliação",
                y_axis_title="Quantidade de Alunos",
                y_axis_min=y_axis_min,
                y_axis_max=y_axis_max,
                output_path=temp_path
            )
        
        if image_path and os.path.exists(image_path):
            try:
                img = Image(image_path)
                # Redimensionar imagem para caber melhor no Excel
                img.width = 600  # Largura em pixels (reduzida)
                img.height = 300  # Altura em pixels (reduzida)
                img.anchor = f"A{chart_row}"
                sheet.add_image(img)
            except Exception as e:
                logger.error(f"Erro ao adicionar imagem do gráfico de níveis: {str(e)}")
        
        ExcelFormatter.auto_adjust_column_width(sheet)
    
    def _create_metric_table(self, sheet, row_start: int, title: str, values: List[float],
                           variations: List[float], directions: List[str],
                           evaluations: List[Dict], y_axis_max: float, y_axis_min: float):
        """Cria tabela de métrica com gráfico - sempre 3 avaliações"""
        col = 1
        row = row_start
        
        # Garantir sempre 3 avaliações
        eval_list = evaluations[:3] if len(evaluations) >= 3 else evaluations
        while len(eval_list) < 3:
            eval_list.append({'title': f'Avaliação {len(eval_list) + 1}', 'id': f'fake_{len(eval_list) + 1}'})
        
        # Garantir sempre 3 valores e 2 variações
        values_list = values[:3] if len(values) >= 3 else values
        while len(values_list) < 3:
            values_list.append(None)
        
        variations_list = variations[:2] if len(variations) >= 2 else variations
        while len(variations_list) < 2:
            variations_list.append(0)
        
        directions_list = directions[:2] if len(directions) >= 2 else directions
        while len(directions_list) < 2:
            directions_list.append('stable')
        
        # Título
        sheet.cell(row, col).value = title
        ExcelFormatter.format_subtitle_cell(sheet.cell(row, col))
        sheet.merge_cells(f"A{row}:{get_column_letter(1 + 3 + 2 + 1)}{row}")  # 3 avaliações + 2 variações + direção
        
        # Cabeçalho
        row += 2
        sheet.cell(row, col).value = "Avaliação"
        ExcelFormatter.format_header_cell(sheet.cell(row, col))
        
        # Colunas de valores (sempre 3)
        data_start_col = col + 1
        for i in range(3):
            col += 1
            eval_title = eval_list[i].get('title', f'Avaliação {i+1}')[:15] if i < len(eval_list) else f'Avaliação {i+1}'
            sheet.cell(row, col).value = eval_title
            ExcelFormatter.format_header_cell(sheet.cell(row, col))
        
        data_end_col = col
        
        # Colunas de variações (sempre 2)
        var_start_col = col + 1
        for i in range(2):
            col += 1
            sheet.cell(row, col).value = f"Var {i+1}→{i+2} (%)"
            ExcelFormatter.format_header_cell(sheet.cell(row, col))
        
        col += 1
        sheet.cell(row, col).value = "Direção"
        ExcelFormatter.format_header_cell(sheet.cell(row, col))
        
        # Linha de valores
        row += 1
        col = 1
        sheet.cell(row, col).value = "Valor"
        ExcelFormatter.format_data_cell(sheet.cell(row, col))
        
        # Contar quantas colunas realmente têm dados (não None)
        actual_data_cols = 0
        for i, value in enumerate(values_list):
            col += 1
            # Se value é None, colocar 0 para evitar problemas no gráfico
            if value is None:
                ExcelFormatter.format_number_cell(sheet.cell(row, col), 0)
                # Contar como coluna válida se não for a última
                if i < len(values_list) - 1:
                    actual_data_cols = i + 1
            else:
                ExcelFormatter.format_number_cell(sheet.cell(row, col), value)
                actual_data_cols = i + 1  # Última coluna com dados
        
        # Garantir pelo menos 2 colunas para ter um gráfico válido
        if actual_data_cols < 2:
            actual_data_cols = len(values_list)  # Usar todas as colunas
        
        # Linha de variações
        var_row = row + 1
        var_col = var_start_col
        for i, (var_value, direction) in enumerate(zip(variations_list, directions_list)):
            ExcelFormatter.format_percentage_cell(sheet.cell(var_row, var_col), var_value, direction)
            var_col += 1
        
        # Direção (primeira variação)
        if directions_list:
            ExcelFormatter.format_direction_cell(sheet.cell(var_row, var_start_col + len(variations_list)), directions_list[0])
        
        # Gráfico - sempre Composed Chart
        # Posicionar gráfico logo abaixo da tabela (5 linhas de espaço)
        chart_row = var_row + 5  # var_row é a última linha da tabela (variações)
        
        # Validar se há dados suficientes para criar gráfico
        if actual_data_cols == 0:
            # Não há dados, pular criação do gráfico
            return var_row + 5  # Retornar posição da tabela se não houver gráfico
        
        # Range de categorias (nomes das avaliações no cabeçalho)
        header_row = row - 1
        actual_end_col = data_start_col + actual_data_cols - 1
        categories_range = f"{get_column_letter(data_start_col)}{header_row}:{get_column_letter(actual_end_col)}{header_row}"
        
        # Range de dados de barras (valores) - apenas colunas com dados
        bar_data_range = f"{get_column_letter(data_start_col)}{row}:{get_column_letter(actual_end_col)}{row}"
        
        # Range de dados de linhas - A linha deve conectar os topos das barras
        # Para isso, a linha deve usar os MESMOS valores das barras (não as variações percentuais)
        # A variação percentual será mostrada como label na linha
        valid_variations = [v for v in variations_list if v is not None]
        valid_directions = [d for i, d in enumerate(directions_list) if i < len(valid_variations)]
        
        # Criar linha sempre que houver pelo menos 2 barras (valores)
        # A linha usa os valores das barras para conectar os topos
        if actual_data_cols >= 2:
            # A linha deve usar os valores das barras (mesmo range)
            line_data_range = bar_data_range  # Usar os mesmos valores das barras
        else:
            line_data_range = None
        
        # Criar gráfico usando Matplotlib (imagem PNG)
        
        # Preparar labels das avaliações
        eval_labels = [eval_list[i].get('title', f'Avaliação {i+1}')[:30] if i < len(eval_list) else f'Avaliação {i+1}' 
                      for i in range(actual_data_cols)]
        
        # Criar arquivo temporário para a imagem
        temp_fd, temp_path = tempfile.mkstemp(suffix='.png', prefix=f'chart_{title.replace(" ", "_")}_')
        os.close(temp_fd)  # Fechar o file descriptor, vamos usar apenas o path
        self.temp_files.append(temp_path)
        
        image_path = None
        if actual_data_cols >= 2 and valid_variations:
            # Criar gráfico combinado com linha conectando os topos das barras
            image_path = MatplotlibChartBuilder.create_composed_chart_image(
                labels=eval_labels,
                values=values_list[:actual_data_cols],
                variations=valid_variations,
                directions=valid_directions,
                chart_title=title,
                y_axis_title="Valor",
                y_axis_min=y_axis_min,
                y_axis_max=y_axis_max,
                output_path=temp_path
            )
        else:
            # Se não houver dados suficientes, criar apenas gráfico de barras
            image_path = MatplotlibChartBuilder.create_bar_chart_image(
                labels=eval_labels,
                values=values_list[:actual_data_cols],
                chart_title=title,
                y_axis_title="Valor",
                y_axis_min=y_axis_min,
                y_axis_max=y_axis_max,
                output_path=temp_path
            )
        
        if image_path and os.path.exists(image_path):
            try:
                img = Image(image_path)
                # Redimensionar imagem para caber melhor no Excel
                img.width = 600  # Largura em pixels (reduzida)
                img.height = 300  # Altura em pixels (reduzida)
                img.anchor = f"A{chart_row}"
                sheet.add_image(img)
                # Retornar a linha final estimada do gráfico (imagens têm aproximadamente 15 linhas de altura)
                # Adicionar espaço grande depois do gráfico
                estimated_chart_end_row = chart_row + 15
                return estimated_chart_end_row + 10  # Retornar linha final + 10 linhas de espaço
            except Exception as e:
                logger.error(f"Erro ao adicionar imagem do gráfico: {str(e)}")
                return var_row + 5
        else:
            return var_row + 5  # Retornar posição da tabela se não houver gráfico

