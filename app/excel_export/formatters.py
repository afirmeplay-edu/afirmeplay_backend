# -*- coding: utf-8 -*-
"""
Formatação de células e estilos para Excel
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Optional, Tuple

# Cores padrão do sistema
COLOR_PRIMARY = "7C3AED"  # Roxo
COLOR_SUCCESS = "00C853"  # Verde
COLOR_DANGER = "FF1744"   # Vermelho
COLOR_WARNING = "FFA726"  # Laranja
COLOR_GRAY = "9E9E9E"     # Cinza
COLOR_LIGHT_GRAY = "F5F5F5"  # Cinza claro
COLOR_WHITE = "FFFFFF"
COLOR_BLACK = "000000"


class ExcelFormatter:
    """Classe para formatação de células Excel"""
    
    # Estilos pré-definidos
    HEADER_FILL = PatternFill(start_color=COLOR_PRIMARY, end_color=COLOR_PRIMARY, fill_type="solid")
    HEADER_FONT = Font(name="Arial", size=12, bold=True, color=COLOR_WHITE)
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    TITLE_FONT = Font(name="Arial", size=16, bold=True, color=COLOR_BLACK)
    SUBTITLE_FONT = Font(name="Arial", size=14, bold=True, color=COLOR_PRIMARY)
    NORMAL_FONT = Font(name="Arial", size=11, color=COLOR_BLACK)
    BOLD_FONT = Font(name="Arial", size=11, bold=True, color=COLOR_BLACK)
    
    BORDER_THIN = Border(
        left=Side(style='thin', color=COLOR_BLACK),
        right=Side(style='thin', color=COLOR_BLACK),
        top=Side(style='thin', color=COLOR_BLACK),
        bottom=Side(style='thin', color=COLOR_BLACK)
    )
    
    ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
    ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
    ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
    
    @staticmethod
    def format_header_cell(cell):
        """Formata célula de cabeçalho"""
        cell.fill = ExcelFormatter.HEADER_FILL
        cell.font = ExcelFormatter.HEADER_FONT
        cell.alignment = ExcelFormatter.HEADER_ALIGNMENT
        cell.border = ExcelFormatter.BORDER_THIN
    
    @staticmethod
    def format_title_cell(cell):
        """Formata célula de título"""
        cell.font = ExcelFormatter.TITLE_FONT
        cell.alignment = ExcelFormatter.ALIGN_CENTER
    
    @staticmethod
    def format_subtitle_cell(cell):
        """Formata célula de subtítulo"""
        cell.font = ExcelFormatter.SUBTITLE_FONT
        cell.alignment = ExcelFormatter.ALIGN_LEFT
    
    @staticmethod
    def format_data_cell(cell, value=None, direction: Optional[str] = None):
        """
        Formata célula de dados com cores baseadas na direção
        
        Args:
            cell: Célula Excel
            value: Valor a ser formatado
            direction: 'increase', 'decrease', 'stable' ou None
        """
        cell.font = ExcelFormatter.NORMAL_FONT
        cell.alignment = ExcelFormatter.ALIGN_CENTER
        cell.border = ExcelFormatter.BORDER_THIN
        
        if direction == 'increase':
            cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")  # Verde claro
        elif direction == 'decrease':
            cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")  # Vermelho claro
        elif direction == 'stable':
            cell.fill = PatternFill(start_color=COLOR_LIGHT_GRAY, end_color=COLOR_LIGHT_GRAY, fill_type="solid")
        else:
            cell.fill = PatternFill(start_color=COLOR_WHITE, end_color=COLOR_WHITE, fill_type="solid")
    
    @staticmethod
    def format_number_cell(cell, value, decimals: int = 2):
        """Formata célula numérica"""
        ExcelFormatter.format_data_cell(cell)
        if value is not None:
            cell.value = round(value, decimals) if isinstance(value, (int, float)) else value
        else:
            cell.value = "-"
    
    @staticmethod
    def format_percentage_cell(cell, value, direction: Optional[str] = None):
        """Formata célula de percentual"""
        ExcelFormatter.format_data_cell(cell, direction=direction)
        if value is not None:
            cell.value = round(value, 2)
            cell.number_format = '0.00"%"'
        else:
            cell.value = "-"
    
    @staticmethod
    def format_direction_cell(cell, direction: Optional[str]):
        """Formata célula de direção (Aumento/Queda/Estável)"""
        ExcelFormatter.format_data_cell(cell, direction=direction)
        
        if direction == 'increase':
            cell.value = "Aumento"
        elif direction == 'decrease':
            cell.value = "Queda"
        elif direction == 'stable':
            cell.value = "Estável"
        else:
            cell.value = "-"
    
    @staticmethod
    def auto_adjust_column_width(worksheet, min_width: int = 10, max_width: int = 50):
        """Ajusta automaticamente a largura das colunas"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max(min_width + 2, max_length + 2), max_width)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def merge_and_center(worksheet, cell_range: str, value: str, font=None, fill=None):
        """Mescla células e centraliza"""
        worksheet.merge_cells(cell_range)
        cell = worksheet[cell_range.split(':')[0]]
        cell.value = value
        cell.alignment = ExcelFormatter.ALIGN_CENTER
        
        if font:
            cell.font = font
        if fill:
            cell.fill = fill

