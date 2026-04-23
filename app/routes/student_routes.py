from flask import Blueprint, request, jsonify, Response, make_response
from app.models.student import Student
from app.models.user import User, RoleEnum
from app.models.school import School
from app.models.studentPasswordLog import StudentPasswordLog
from sqlalchemy.exc import SQLAlchemyError, IntegrityError
import logging
from app.decorators.role_required import role_required, get_current_user_from_token
from app.decorators import requires_city_context
from app.decorators.role_required import get_current_tenant_id
from datetime import datetime
from app import db
from app.multitenant.flask_g import get_orm_session
from flask_jwt_extended import jwt_required
from app.utils.auth import hash_password
from marshmallow import ValidationError
from app.models.studentAnswer import StudentAnswer
from app.models.studentClass import Class
from app.models.grades import Grade
from app.models.classTest import ClassTest
from app.models.city import City
from app.models.manager import Manager
from app.models.teacher import Teacher
from app.models.schoolTeacher import SchoolTeacher
from app.models.teacherClass import TeacherClass
from sqlalchemy.orm import joinedload
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import os
import shutil
import tempfile
from jinja2 import Environment, FileSystemLoader, select_autoescape
from weasyprint import HTML

bp = Blueprint('students', __name__, url_prefix="/students")


def _get_logo_path():
    """Retorna o caminho absoluto da logo do app (app/assets/afirme_logo.png) ou None."""
    app_dir = os.path.dirname(os.path.dirname(__file__))
    logo_path = os.path.join(app_dir, 'assets', 'afirme_logo.png')
    return logo_path if os.path.exists(logo_path) else None

@bp.errorhandler(SQLAlchemyError)
def handle_db_error(error):
    get_orm_session().rollback()
    logging.error(f"Database error: {str(error)}")
    return jsonify({"error": "Database error occurred", "details": str(error)}), 500

@bp.errorhandler(ValidationError)
def handle_validation_error(error):
    return jsonify({"error": "Validation error", "details": error.messages}), 400

@bp.errorhandler(Exception)
def handle_generic_error(error):
    logging.error(f"Unexpected error: {str(error)}", exc_info=True)
    return jsonify({"error": "An unexpected error occurred", "details": str(error)}), 500


@bp.route("", methods=['POST'])
@jwt_required()
@role_required("admin", "professor", "coordenador", "diretor", "tecadm")
@requires_city_context
def criar_usuario_e_aluno():
    try:
        logging.info("Iniciando criação de usuário/aluno combinada")

        data = request.get_json()
        logging.info(f"Dados recebidos: {data}")

        if not data:
            return jsonify({"error": "No data provided"}), 400

        required_fields = ["name", "email", "password", "class_id"]
        for field in required_fields:
            if field not in data:
                return jsonify({"error": f"Missing required field: {field}"}), 400

        # Buscar turma para determinar city_id (necessário tanto para usuário novo quanto existente)
        class_obj = get_orm_session().query(Class).get(data["class_id"])
        if not class_obj:
            return jsonify({"error": "Class not found"}), 404
        
        # Buscar escola da turma
        # ✅ CORRIGIDO: Converter para string (School.id é VARCHAR)
        from app.utils.uuid_helpers import uuid_to_str
        school_id_str = uuid_to_str(class_obj.school_id)
        school = get_orm_session().query(School).filter(School.id == school_id_str).first() if school_id_str else None
        if not school:
            return jsonify({"error": "School not found for the specified class"}), 404
        
        # Determinar city_id baseado na escola
        city_id = school.city_id

        # Tenta encontrar o usuário pelo email
        usuario = get_orm_session().query(User).filter_by(email=data["email"]).first()
        usuario_foi_criado_agora = False  # Flag para saber se criamos o usuário agora

        if usuario:
            logging.info(f"Usuário existente encontrado: {usuario.email}")
            # Verificar se o usuário já é um aluno
            if get_orm_session().query(Student).filter_by(user_id=usuario.id).first():
                logging.warning(f"Usuário {usuario.email} já é um aluno.") 
                return jsonify({"error": "User is already a student"}), 400
        else:
            logging.info("Usuário não encontrado, criando novo usuário.")
            usuario_foi_criado_agora = True  # Marcar que vamos criar o usuário agora
            # Verificar se matrícula já existe (caso fornecida) - apenas para novo usuário
            if data.get("registration") and get_orm_session().query(User).filter_by(registration=data["registration"]).first():
                return jsonify({"error": "Registration number already exists"}), 400
            
            # Criar usuário (role padrão: aluno) com city_id da escola
            usuario = User(
                name=data["name"],
                email=data["email"],
                password_hash=hash_password(data["password"]),
                registration=data.get("registration"),
                role=RoleEnum("aluno"),
                city_id=city_id
            )
            get_orm_session().add(usuario)
            get_orm_session().flush() 
            print(usuario.id)
            logging.info(f"Novo usuário criado com sucesso. ID: {usuario.id}")

        # Verificar formato da data de nascimento (se fornecida)
        birth_date = None
        if "birth_date" in data:
            try:
                birth_date = datetime.strptime(data["birth_date"], "%Y-%m-%d").date()
            except ValueError:
                # Se o usuário foi criado neste request, precisamos desfazê-lo
                if not usuario.id:
                    get_orm_session().rollback()
                return jsonify({"error": "Invalid birth date format. Use YYYY-MM-DD"}), 400

        # Determinar grade_id (série) com fallback da turma
        grade_id = data.get("grade_id") or class_obj.grade_id

        # Criar aluno
        novo_aluno = Student(
            name=usuario.name,
            user_id=usuario.id,
            registration=usuario.registration,
            birth_date=birth_date,
            class_id=data["class_id"],
            grade_id=grade_id,
            school_id=class_obj.school_id
        )
        get_orm_session().add(novo_aluno)
        get_orm_session().flush()  # Flush para obter o ID do aluno
        
        # Salvar senha em texto plano na tabela de log apenas se criamos um novo usuário
        # (quando o usuário já existe, não temos a senha original em texto plano)
        if usuario_foi_criado_agora:
            password_log = StudentPasswordLog(
                student_name=data["name"],
                email=data["email"],
                password=data["password"],  # Senha em texto plano
                registration=data.get("registration"),
                user_id=usuario.id,
                student_id=novo_aluno.id,
                class_id=data["class_id"],
                grade_id=grade_id,
                school_id=class_obj.school_id,
                city_id=city_id
            )
            get_orm_session().add(password_log)
        
        get_orm_session().commit()

        logging.info(f"Aluno criado com sucesso para o usuário ID: {usuario.id}")

        return jsonify({
            "message": "Student and/or User created successfully!",
            "user": {
                "id": usuario.id,
                "name": usuario.name,
                "email": usuario.email,
                "registration": usuario.registration,
                "role": usuario.role.value,
                "city_id": usuario.city_id
            },
            "student": {
                "id": novo_aluno.id,
                "name": novo_aluno.name,
                "registration": novo_aluno.registration,
                "birth_date": str(novo_aluno.birth_date) if novo_aluno.birth_date else None,
                "class_id": novo_aluno.class_id,
                "grade_id": str(novo_aluno.grade_id) if novo_aluno.grade_id else None,
                "school_id": novo_aluno.school_id
            }
        }), 201

    except SQLAlchemyError as e:
        get_orm_session().rollback()
        logging.error(f"Database error during user/student creation: {str(e)}")
        return jsonify({"error": "Database error occurred", "details": str(e)}), 500
    except Exception as e:
        get_orm_session().rollback()
        logging.error(f"Unexpected error during user/student creation: {str(e)}", exc_info=True)
        return jsonify({"error": "Unexpected error occurred", "details": str(e)}), 500


# Helper function to format student data with related details
def format_student_details(student, user=None, school=None, class_obj=None, grade=None):
    user_details = None
    if user:
        user_details = {
            "id": user.id,
            "name": user.name,
            "email": user.email,
            "registration": user.registration,
            "role": user.role.value,
            "city_id": user.city_id,
            "created_at": user.created_at.isoformat() if user.created_at else None,
            "updated_at": user.updated_at.isoformat() if user.updated_at else None
        }

    school_details = None
    if school:
        school_details = {
            "id": school.id,
            "name": school.name,
            "domain": school.domain,
            "address": school.address,
            "city_id": school.city_id,
            "created_at": school.created_at.isoformat() if school.created_at else None
        }

    class_details = None
    if class_obj:
        class_details = {
            "id": class_obj.id,
            "name": class_obj.name,
            "school_id": class_obj.school_id,
            "grade_id": str(class_obj.grade_id) if class_obj.grade_id else None
        }

    grade_details = None
    if grade:
        grade_details = {
            "id": grade.id,
            "name": grade.name
        }

    return {
        "id": student.id,
        "name": student.name,
        "registration": student.registration,
        "birth_date": student.birth_date.isoformat() if student.birth_date else None,
        "class_id": student.class_id,
        "grade_id": student.grade_id,
        "school_id": student.school_id,
        "created_at": student.created_at.isoformat() if student.created_at else None,
        "user": user_details,
        "school": school_details,
        "class": class_details,
        "grade": grade_details
    }

# GET - Listar alunos
@bp.route('', methods=['GET'])
@jwt_required()
@role_required("admin", "diretor", "coordenador", "professor", "tecadm")
def listar_alunos():
    try:
        user = get_current_user_from_token()

        if not user:
            return jsonify({"message": "Usuário não encontrado"}), 404

        # Base query with joins - usando LEFT JOIN para incluir alunos sem escola
        query = get_orm_session().query(
            Student,
            User,
            School,
            Class,
            Grade
        ).join(
            User, Student.user_id == User.id
        ).outerjoin(
            School, Student.school_id == School.id
        ).outerjoin(
            Class, Student.class_id == Class.id
        ).outerjoin(
            Grade, Student.grade_id == Grade.id
        )

        students = []
        if user['role'] == "admin":
            # Admin pode ver todos os alunos
            students = query.all()
        elif user['role'] == "professor":
            # Professor só vê alunos das escolas onde está vinculado
            from app.models.teacher import Teacher
            from app.models.schoolTeacher import SchoolTeacher
            
            teacher = get_orm_session().query(Teacher).filter_by(user_id=user['id']).first()
            if not teacher:
                return jsonify({"message": "Professor não encontrado"}), 404
            
            # Buscar escolas onde o professor está vinculado
            teacher_schools = get_orm_session().query(SchoolTeacher).filter_by(teacher_id=teacher.id).all()
            school_ids = [ts.school_id for ts in teacher_schools]
            
            if school_ids:
                students = query.filter(Student.school_id.in_(school_ids)).all()
            else:
                return jsonify({"message": "Professor não está alocado em nenhuma escola"}), 400
        elif user['role'] in ["diretor", "coordenador"]:
            # Diretor e coordenador veem alunos apenas de sua escola (via Manager, não Teacher)
            from app.models.manager import Manager
            from app.utils.uuid_helpers import uuid_to_str

            manager = get_orm_session().query(Manager).filter_by(user_id=user['id']).first()
            if not manager:
                return jsonify({"message": "Diretor/Coordenador não encontrado na tabela manager"}), 404
            school_id_str = uuid_to_str(manager.school_id) if manager.school_id else None
            if not school_id_str:
                return jsonify({"message": "Diretor/Coordenador não está vinculado a nenhuma escola"}), 400

            students = query.filter(Student.school_id == school_id_str).all()
        else:
            # TecAdmin vê alunos de todas as escolas do município + alunos sem escola mas com city_id
            city_id = user.get('tenant_id') or user.get('city_id')
            if not city_id:
                return jsonify({"message": "ID da cidade não disponível para este usuário"}), 400

            # Busca todas as escolas da cidade
            schools_in_city = get_orm_session().query(School).filter_by(city_id=city_id).all()
            school_ids = [school.id for school in schools_in_city]

            # Busca alunos das escolas da cidade OU alunos sem escola mas com city_id correto
            students = query.filter(
                db.or_(
                    Student.school_id.in_(school_ids),
                    db.and_(Student.school_id.is_(None), User.city_id == city_id)
                )
            ).all()

        return jsonify([format_student_details(student, user, school, class_obj, grade) 
                       for student, user, school, class_obj, grade in students]), 200

    except SQLAlchemyError as e:
        logging.error(f"Erro no banco de dados ao listar alunos: {e}")
        return jsonify({"message": "Erro interno do servidor ao consultar dados", "details": str(e)}), 500
    except AttributeError as e:
        logging.error(f"Erro de atributo ao processar usuário ou papel: {e}")
        return jsonify({"message": "Erro ao processar dados do usuário", "details": str(e)}), 500
    except Exception as e:
        logging.error(f"Erro inesperado na rota listar_alunos: {e}", exc_info=True)
        return jsonify({"message": "Ocorreu um erro inesperado no servidor"}), 500

# GET - Obter dados completos de um aluno específico
@bp.route('/<string:student_id>', methods=['GET'])
@jwt_required()
@role_required("admin", "diretor", "coordenador", "professor", "tecadm", "aluno")
def obter_aluno_completo(student_id):
    """
    Retorna todos os dados do aluno incluindo:
    - Nome completo
    - Escola
    - Estado e Município
    - Turma atrelada
    - Série
    - Professores vinculados à turma
    """
    try:
        user = get_current_user_from_token()
        if not user:
            return jsonify({"message": "Usuário não encontrado"}), 404

        # Buscar o aluno com todas as relações
        # IMPORTANTE: student_id é na verdade o user_id (ID da tabela users)
        student = get_orm_session().query(
            Student,
            User,
            School,
            Class,
            Grade
        ).join(
            User, Student.user_id == User.id
        ).outerjoin(
            School, Student.school_id == School.id
        ).outerjoin(
            Class, Student.class_id == Class.id
        ).outerjoin(
            Grade, Student.grade_id == Grade.id
        ).filter(
            Student.user_id == student_id
        ).first()

        if not student:
            return jsonify({"message": "Aluno não encontrado"}), 404

        student_obj, user_obj, school, class_obj, grade = student

        # Verificar permissões
        if user['role'] == "aluno":
            # Aluno só pode ver seus próprios dados
            if student_obj.user_id != user['id']:
                return jsonify({"message": "Você não tem permissão para visualizar este aluno"}), 403
        elif user['role'] == "professor":
            # Professor só pode ver alunos das escolas onde está vinculado
            from app.models.teacher import Teacher
            from app.models.schoolTeacher import SchoolTeacher
            
            teacher = get_orm_session().query(Teacher).filter_by(user_id=user['id']).first()
            if not teacher:
                return jsonify({"message": "Professor não encontrado"}), 404
            
            teacher_schools = get_orm_session().query(SchoolTeacher).filter_by(teacher_id=teacher.id).all()
            school_ids = [ts.school_id for ts in teacher_schools]
            
            if not school_ids or (student_obj.school_id and student_obj.school_id not in school_ids):
                return jsonify({"message": "Você não tem permissão para visualizar este aluno"}), 403
        elif user['role'] in ["diretor", "coordenador"]:
            # Diretor e coordenador só podem ver alunos de sua escola (via Manager)
            from app.models.manager import Manager
            from app.utils.uuid_helpers import uuid_to_str

            manager = get_orm_session().query(Manager).filter_by(user_id=user['id']).first()
            if not manager:
                return jsonify({"message": "Diretor/Coordenador não encontrado"}), 404
            manager_school_id_str = uuid_to_str(manager.school_id) if manager.school_id else None
            if not manager_school_id_str or student_obj.school_id != manager_school_id_str:
                return jsonify({"message": "Você não tem permissão para visualizar este aluno"}), 403
        elif user['role'] == "tecadm":
            # TecAdmin só pode ver alunos do seu município
            city_id = user.get('tenant_id') or user.get('city_id')
            if not city_id:
                return jsonify({"message": "ID da cidade não disponível para este usuário"}), 400
            
            if school and school.city_id != city_id:
                return jsonify({"message": "Você não tem permissão para visualizar este aluno"}), 403

        # Buscar município e estado através da escola
        city = None
        if school and school.city_id:
            city = get_orm_session().query(City).get(school.city_id)

        # Buscar professores vinculados à turma do aluno
        teachers_data = []
        if class_obj:
            teacher_classes = get_orm_session().query(TeacherClass).filter_by(class_id=class_obj.id).all()
            
            for tc in teacher_classes:
                teacher = get_orm_session().query(Teacher).get(tc.teacher_id)
                if teacher:
                    teacher_user = get_orm_session().query(User).get(teacher.user_id)
                    teachers_data.append({
                        "id": teacher.id,
                        "name": teacher.name,
                        "email": teacher_user.email if teacher_user else None,
                        "registration": teacher.registration,
                        "user_id": teacher.user_id
                    })

        # Montar resposta completa
        response_data = {
            "id": student_obj.id,
            "name": student_obj.name or user_obj.name,
            "full_name": user_obj.name,  # Nome completo do usuário
            "email": user_obj.email,
            "registration": student_obj.registration or user_obj.registration,
            "birth_date": student_obj.birth_date.isoformat() if student_obj.birth_date else None,
            "address": user_obj.address,  # Endereço do aluno
            "created_at": student_obj.created_at.isoformat() if student_obj.created_at else None,
            "school": {
                "id": school.id,
                "name": school.name,
                "address": school.address,
                "domain": school.domain
            } if school else None,
            "city": {
                "id": city.id,
                "name": city.name,
                "state": city.state
            } if city else None,
            "municipio": city.name if city else None,  # Nome do município
            "estado": city.state if city else None,  # Estado
            "class": {
                "id": class_obj.id,
                "name": class_obj.name,
                "school_id": class_obj.school_id
            } if class_obj else None,
            "turma": {
                "id": class_obj.id,
                "name": class_obj.name,
                "school_id": class_obj.school_id
            } if class_obj else None,  # Turma atrelada
            "grade": {
                "id": str(grade.id),
                "name": grade.name
            } if grade else None,
            "serie": {
                "id": str(grade.id),
                "name": grade.name
            } if grade else None,  # Série
            "teachers": teachers_data,  # Professores vinculados à turma
            "professores": teachers_data  # Alias para professores
        }

        return jsonify(response_data), 200

    except SQLAlchemyError as e:
        logging.error(f"Erro no banco de dados ao obter aluno completo: {str(e)}")
        return jsonify({"message": "Erro interno do servidor ao consultar dados", "details": str(e)}), 500
    except Exception as e:
        logging.error(f"Erro inesperado na rota obter_aluno_completo: {str(e)}", exc_info=True)
        return jsonify({"message": "Ocorreu um erro inesperado no servidor", "details": str(e)}), 500

@bp.route('/<string:student_id>/<string:class_id>', methods=['PUT'])
@jwt_required()
@role_required("admin", "professor", "coordenador", "diretor")
def atualizar_aluno(student_id, class_id):
    try:
        aluno = get_orm_session().query(Student).filter_by(id=student_id, class_id=class_id).first()

        if not aluno:
            return jsonify({"error": "Student not found"}), 404

        dados = request.get_json()
        if not dados:
            return jsonify({"error": "No data provided"}), 400

        # Atualiza dados do usuário vinculado
        usuario = aluno.user
        if not usuario:
            return jsonify({"error": "Associated user not found"}), 404

        # Armazenar valores antigos para tracking
        old_class_id = aluno.class_id
        old_school_id = aluno.school_id
        class_changed = False
        school_changed = False
        
        # Processar mudança de turma
        if "class_id" in dados and dados["class_id"] != old_class_id:
            from app.utils.uuid_helpers import ensure_uuid
            new_class_id = ensure_uuid(dados["class_id"])
            
            if not new_class_id:
                return jsonify({"error": "ID de turma inválido"}), 400
            
            # Buscar nova turma
            from app.models.studentClass import Class
            new_class = get_orm_session().query(Class).get(new_class_id)
            
            if not new_class:
                return jsonify({"error": "Nova turma não encontrada"}), 404
            
            # Atualizar turma
            aluno.class_id = new_class_id
            class_changed = True
            
            # Se a turma pertence a outra escola, atualizar automaticamente
            if new_class.school_id != old_school_id:
                aluno.school_id = new_class.school_id
                school_changed = True
                
                # Atualizar city_id do usuário
                new_school = get_orm_session().query(School).get(new_class.school_id)
                if new_school and new_school.city_id != usuario.city_id:
                    usuario.city_id = new_school.city_id
                    logging.info(f"Aluno {student_id} movido para escola {new_class.school_id}. City_id atualizado para {new_school.city_id}")
        
        # Atualizar outros dados do usuário
        if "name" in dados:
            usuario.name = dados["name"]
        if "email" in dados:
            usuario.email = dados["email"]
        if "registration" in dados:
            usuario.registration = dados["registration"]

        # Atualizar série
        if "grade_id" in dados:
            aluno.grade_id = dados["grade_id"]

        # Atualiza data de nascimento se enviada
        if "birth_date" in dados:
            try:
                aluno.birth_date = datetime.strptime(dados["birth_date"], "%Y-%m-%d").date()
            except ValueError:
                return jsonify({"error": "Invalid date format. Use YYYY-MM-DD"}), 400

        # Se school_id mudou mas class_id não (caso raro), atualizar city_id
        if not class_changed and aluno.school_id:
            school = get_orm_session().query(School).get(aluno.school_id)
            if school and school.city_id != usuario.city_id:
                usuario.city_id = school.city_id
                logging.info(f"Atualizando city_id do aluno {usuario.id} para {school.city_id}")

        get_orm_session().commit()
        
        # Preparar resposta detalhada
        response = {
            "message": "Aluno atualizado com sucesso",
            "student": {
                "id": aluno.id,
                "name": usuario.name,
                "email": usuario.email,
                "class_id": str(aluno.class_id) if aluno.class_id else None,
                "school_id": aluno.school_id,
                "grade_id": str(aluno.grade_id) if aluno.grade_id else None
            },
            "changes": {
                "class_changed": class_changed,
                "school_changed": school_changed
            }
        }
        
        if class_changed or school_changed:
            response["relocation"] = {
                "old_class_id": str(old_class_id) if old_class_id else None,
                "new_class_id": str(aluno.class_id) if aluno.class_id else None,
                "old_school_id": old_school_id,
                "new_school_id": aluno.school_id,
                "city_updated": school_changed
            }
        
        return jsonify(response), 200

    except IntegrityError as e:
        get_orm_session().rollback()
        return jsonify({"error": "Database integrity error", "details": str(e)}), 400
    except Exception as e:
        get_orm_session().rollback()
        logging.error(f"Error updating student: {str(e)}", exc_info=True)
        return jsonify({"error": "Error updating student", "details": str(e)}), 500

@bp.route('/<string:aluno_id>', methods=['DELETE'])
@jwt_required()
@role_required("admin", "professor", "coordenador", "diretor")
def deletar_aluno(aluno_id):
    try:
        aluno = get_orm_session().query(Student).filter_by(id=aluno_id).first()

        if not aluno:
            return jsonify({"error": "Student not found"}), 404

        # Verifica se existe usuário associado
        if aluno.user:
            get_orm_session().delete(aluno.user)
        
        get_orm_session().delete(aluno)
        get_orm_session().commit()
        return jsonify({"message": "Student deleted successfully"}), 200

    except IntegrityError as e:
        get_orm_session().rollback()
        return jsonify({"error": "Database integrity error", "details": str(e)}), 400
    except Exception as e:
        get_orm_session().rollback()
        logging.error(f"Error deleting student: {str(e)}", exc_info=True)
        return jsonify({"error": "Error deleting student", "details": str(e)}), 500

@bp.route('/available', methods=['GET'])
@jwt_required()
@role_required("admin", "diretor", "coordenador", "professor", "tecadm")
def get_available_students():
    """
    Lista alunos que não estão vinculados a nenhuma escola (disponíveis para alocação)
    """
    try:
        user = get_current_user_from_token()
        if not user:
            return jsonify({"message": "Usuário não encontrado"}), 404

        # Buscar TODOS os usuários com role aluno (disponíveis para alocação)
        # Isso inclui tanto usuários sem registro Student quanto com Student mas sem school_id
        all_users_with_aluno_role = get_orm_session().query(
            User
        ).filter(
            User.role == RoleEnum.ALUNO
        ).all()
        
        logging.info(f"Todos os usuários com role aluno: {[(u.id, u.name, u.city_id) for u in all_users_with_aluno_role]}")
        
        # Filtrar apenas os que são realmente "disponíveis" (sem escola)
        available_students = []
        
        for user_obj in all_users_with_aluno_role:
            # Verificar se tem registro na tabela Student
            student_record = get_orm_session().query(Student).filter_by(user_id=user_obj.id).first()
            
            if student_record:
                # Se tem registro Student, só é disponível se school_id for NULL
                if student_record.school_id is None:
                    available_students.append((student_record, user_obj))
                    logging.info(f"Aluno disponível (com Student, sem escola): {user_obj.name} (city_id: {user_obj.city_id})")
            else:
                # Se não tem registro Student, é disponível
                available_students.append((None, user_obj))
                logging.info(f"Aluno disponível (sem Student): {user_obj.name} (city_id: {user_obj.city_id})")
        
        students = available_students
        logging.info(f"Total de alunos disponíveis: {len(students)}")

        # Filtrar por permissões
        logging.info(f"Verificando permissões para role: {user['role']}")
        
        if user['role'] == "admin":
            # Admin vê todos os alunos disponíveis
            logging.info("Admin - permitindo acesso a todos os alunos disponíveis")
            pass
        elif user['role'] == "tecadm":
            # Tecadm vê apenas alunos do seu município
            city_id = user.get('tenant_id') or user.get('city_id')
            logging.info(f"Tecadm - city_id: {city_id}")
            if not city_id:
                return jsonify({"message": "ID da cidade não disponível para este usuário"}), 400
            
            # Log antes do filtro
            print(f"Alunos antes do filtro por city_id: {[(s[0].id if s[0] else 'None', s[1].id, s[1].name, s[1].city_id) for s in students]}")
            logging.info(f"Alunos antes do filtro por city_id: {[(s[0].id if s[0] else 'None', s[1].id, s[1].name, s[1].city_id) for s in students]}")
            
            # Filtro com verificação detalhada
            filtered_students = []
            for s in students:
                user_city_id = s[1].city_id
                logging.info(f"Comparando: user_city_id={user_city_id} (tipo: {type(user_city_id)}) vs city_id={city_id} (tipo: {type(city_id)})")
                if user_city_id == city_id:
                    filtered_students.append(s)
                    logging.info(f"Match encontrado para usuário {s[1].name} (city_id: {user_city_id})")
                else:
                    logging.info(f"Não match para usuário {s[1].name} (city_id: {user_city_id})")
            
            students = filtered_students
            logging.info(f"Tecadm - {len(students)} alunos filtrados por city_id")
            
            # Log após o filtro
            logging.info(f"Alunos após filtro por city_id: {[(s[0].id if s[0] else 'None', s[1].id, s[1].name, s[1].city_id) for s in students]}")
        elif user['role'] in ["diretor", "coordenador"]:
            # Diretor e coordenador podem ver alunos disponíveis do seu município
            city_id = user.get('tenant_id') or user.get('city_id')
            logging.info(f"{user['role'].title()} - city_id: {city_id}")
            if not city_id:
                return jsonify({"message": "ID da cidade não disponível para este usuário"}), 400
            
            # Log antes do filtro
            print(f"Alunos antes do filtro por city_id: {[(s[0].id if s[0] else 'None', s[1].id, s[1].name, s[1].city_id) for s in students]}")
            logging.info(f"Alunos antes do filtro por city_id: {[(s[0].id if s[0] else 'None', s[1].id, s[1].name, s[1].city_id) for s in students]}")
            
            # Filtro com verificação detalhada
            filtered_students = []
            for s in students:
                user_city_id = s[1].city_id
                logging.info(f"Comparando: user_city_id={user_city_id} (tipo: {type(user_city_id)}) vs city_id={city_id} (tipo: {type(city_id)})")
                if user_city_id == city_id:
                    filtered_students.append(s)
                    logging.info(f"Match encontrado para usuário {s[1].name} (city_id: {user_city_id})")
                else:
                    logging.info(f"Não match para usuário {s[1].name} (city_id: {user_city_id})")
            
            students = filtered_students
            print(f"Alunos após filtro por city_id: {[(s[0].id if s[0] else 'None', s[1].id, s[1].name, s[1].city_id) for s in students]}")
            logging.info(f"{user['role'].title()} - {len(students)} alunos filtrados por city_id")
            
            # Log após o filtro
            print(f"Alunos após filtro por city_id: {[(s[0].id if s[0] else 'None', s[1].id, s[1].name, s[1].city_id) for s in students]}")
            logging.info(f"Alunos após filtro por city_id: {[(s[0].id if s[0] else 'None', s[1].id, s[1].name, s[1].city_id) for s in students]}")
        elif user['role'] == "professor":
            # Professor não pode ver alunos disponíveis
            logging.info(f"Role {user['role']} - negando acesso")
            return jsonify({"message": "Você não tem permissão para visualizar alunos disponíveis"}), 403

        result = []
        for student, user_obj in students:
            # Se student é None, usar dados do user_obj
            if student is None:
                result.append({
                    "id": None,  # Não tem ID de student
                    "name": user_obj.name,
                    "registration": user_obj.registration,
                    "birth_date": None,  # Não tem birth_date no user
                    "user": {
                        "id": user_obj.id,
                        "name": user_obj.name,
                        "email": user_obj.email,
                        "registration": user_obj.registration,
                        "role": user_obj.role.value,
                        "city_id": user_obj.city_id
                    }
                })
            else:
                result.append({
                    "id": student.id,
                    "name": student.name,
                    "registration": student.registration,
                    "birth_date": student.birth_date.isoformat() if student.birth_date else None,
                    "user": {
                        "id": user_obj.id,
                        "name": user_obj.name,
                        "email": user_obj.email,
                        "registration": user_obj.registration,
                        "role": user_obj.role.value,
                        "city_id": user_obj.city_id
                    }
                })

        return jsonify(result), 200

    except Exception as e:
        logging.error(f"Erro ao listar alunos disponíveis: {str(e)}", exc_info=True)
        return jsonify({"message": "Erro ao listar alunos disponíveis", "details": str(e)}), 500

@bp.route('/school/<string:school_id>', methods=['GET'])
@jwt_required()
@role_required("admin", "professor", "coordenador", "diretor","tecadm")
def get_students_by_school(school_id):
    try:
        logging.info(f"Fetching students for school_id: {school_id}")
        
        user = get_current_user_from_token()
        if not user:
            logging.error("User not found in token")
            return jsonify({"message": "User not found"}), 404

        # Verify if the school exists
        school = get_orm_session().query(School).filter_by(id=school_id).first()
        if not school:
            logging.error(f"School not found with id: {school_id}")
            return jsonify({"message": "School not found"}), 404

        # Check user permissions
        if user['role'] == "admin":
            # Admin can access any school
            pass
        elif user['role'] == "professor":
            # Professor can only see students from their assigned schools
            from app.models.teacher import Teacher
            from app.models.schoolTeacher import SchoolTeacher
            
            teacher = get_orm_session().query(Teacher).filter_by(user_id=user['id']).first()
            if not teacher:
                return jsonify({"message": "Professor não encontrado"}), 404
            
            teacher_schools = get_orm_session().query(SchoolTeacher).filter_by(teacher_id=teacher.id).all()
            teacher_school_ids = [ts.school_id for ts in teacher_schools]
            
            if school_id not in teacher_school_ids:
                logging.warning(f"Professor {user.get('id')} tried to access students from school {school_id}")
                return jsonify({"message": "You don't have permission to view students from this school"}), 403
        elif user['role'] in ["diretor", "coordenador"]:
            # Director and coordinator can only see students from their school
            from app.models.manager import Manager
            
            manager = get_orm_session().query(Manager).filter_by(user_id=user['id']).first()
            if not manager:
                return jsonify({"message": "Diretor/Coordenador não encontrado na tabela manager"}), 404
            
            if not manager.school_id or manager.school_id != school_id:
                logging.warning(f"User {user.get('id')} tried to access students from school {school_id}")
                return jsonify({"message": "You don't have permission to view students from this school"}), 403
        else:
            # TecAdmin can access schools in their municipality
            city_id = user.get('tenant_id') or user.get('city_id')
            if not city_id or school.city_id != city_id:
                logging.warning(f"User {user.get('id')} tried to access students from school in different city")
                return jsonify({"message": "You don't have permission to view students from this school"}), 403

        # Get all students from the school with related data using joins
        try:
            students = get_orm_session().query(
                Student,
                User,
                School,
                Class,
                Grade
            ).join(
                User, Student.user_id == User.id
            ).join(
                School, Student.school_id == School.id
            ).join(
                Class, Student.class_id == Class.id
            ).outerjoin(
                Grade, Student.grade_id == Grade.id
            ).filter(
                Student.school_id == school_id
            ).all()
            
            logging.info(f"Found {len(students)} students for school {school_id}")
        except SQLAlchemyError as e:
            logging.error(f"Database error while querying students: {str(e)}")
            return jsonify({"message": "Error querying students from database"}), 500

        return jsonify([format_student_details(student, user, school, class_obj, grade) 
                       for student, user, school, class_obj, grade in students]), 200

    except SQLAlchemyError as e:
        logging.error(f"Database error while fetching students by school: {str(e)}")
        return jsonify({"message": "Internal server error while querying data", "details": str(e)}), 500
    except Exception as e:
        logging.error(f"Unexpected error in get_students_by_school route: {str(e)}", exc_info=True)
        return jsonify({"message": "An unexpected error occurred", "details": str(e)}), 500

@bp.route('/classes/<string:class_id>', methods=['GET'])
@jwt_required()
@role_required("admin", "diretor", "coordenador", "professor","tecadm")
def get_students_by_class(class_id):
    try:
        logging.info(f"Fetching students for class_id: {class_id}")

        # Converter class_id para UUID (Class.id é UUID)
        from app.utils.uuid_helpers import ensure_uuid
        class_id_uuid = ensure_uuid(class_id)
        if not class_id_uuid:
            return jsonify({"message": "ID de turma inválido"}), 400
        
        # Check if the class exists (optional, but good practice)
        class_obj = get_orm_session().query(Class).get(class_id_uuid)
        if not class_obj:
            logging.warning(f"Class not found with ID: {class_id}")
            return jsonify({"message": "Class not found"}), 404

        # Query with explicit joins to get all related data
        students = get_orm_session().query(
            Student,
            User,
            School,
            Class,
            Grade
        ).join(
            User, Student.user_id == User.id
        ).join(
            School, Student.school_id == School.id
        ).join(
            Class, Student.class_id == Class.id
        ).outerjoin(
            Grade, Student.grade_id == Grade.id
        ).filter(
            Student.class_id == class_id_uuid
        ).all()

        if not students:
            logging.info(f"No students found for class_id: {class_id}")
            return jsonify([]), 200  # Return empty list if no students found

        return jsonify([format_student_details(student, user, school, class_obj, grade) 
                       for student, user, school, class_obj, grade in students]), 200

    except SQLAlchemyError as e:
        logging.error(f"Database error while fetching students by class: {str(e)}")
        return jsonify({"message": "Internal server error while querying data", "details": str(e)}), 500
    except Exception as e:
        logging.error(f"Unexpected error in get_students_by_class route: {str(e)}", exc_info=True)
        return jsonify({"message": "An unexpected error occurred", "details": str(e)}), 500

@bp.route('/school/<string:school_id>/class/<string:class_id>', methods=['GET'])
@jwt_required()
@role_required("admin", "diretor", "coordenador", "professor","tecadm")
def get_students_by_school_and_class(school_id, class_id):
    try:
        logging.info(f"Fetching students for school_id: {school_id} and class_id: {class_id}")

        # Primeiro validar se escola e turma existem
        school = get_orm_session().query(School).get(school_id)
        if not school:
            logging.warning(f"School not found with id: {school_id}")
            return jsonify({"error": "School not found"}), 404

        # Converter class_id para UUID (Class.id é UUID)
        from app.utils.uuid_helpers import ensure_uuid
        class_id_uuid = ensure_uuid(class_id)
        if not class_id_uuid:
            return jsonify({"error": "ID de turma inválido"}), 400
        
        class_obj = get_orm_session().query(Class).get(class_id_uuid)
        if not class_obj:
            logging.warning(f"Class not found with id: {class_id}")
            return jsonify({"error": "Class not found"}), 404

        # Validar se a turma pertence à escola especificada
        # Converter school_id para UUID (Class.school_id é UUID)
        school_id_uuid = ensure_uuid(school_id)
        if not school_id_uuid:
            return jsonify({"error": "ID de escola inválido"}), 400
        
        if class_obj.school_id != school_id_uuid:
            logging.warning(f"Class {class_id} does not belong to school {school_id}")
            return jsonify({"error": "Class does not belong to the specified school"}), 400

        # Agora buscar os alunos com validação correta
        # Primeiro buscar apenas os alunos para garantir que todos sejam retornados
        # Converter class_id para UUID (Student.class_id é UUID)
        class_id_uuid = ensure_uuid(class_id)
        if not class_id_uuid:
            return jsonify({"error": "ID de turma inválido"}), 400
        
        students = get_orm_session().query(Student).filter_by(
            school_id=school_id,
            class_id=class_id_uuid
        ).all()
        
        if not students:
            logging.info(f"No students found for school_id: {school_id} and class_id: {class_id}")
            return jsonify([]), 200  # Return empty list if no students found
        
        # Agora buscar os dados relacionados para cada aluno
        formatted_students = []
        for student in students:
            user = get_orm_session().query(User).get(student.user_id)
            # ✅ CORRIGIDO: Converter para string (School.id é VARCHAR)
            from app.utils.uuid_helpers import uuid_to_str
            school_id_str = uuid_to_str(student.school_id)
            school = get_orm_session().query(School).filter(School.id == school_id_str).first() if school_id_str else None
            class_obj = get_orm_session().query(Class).get(student.class_id)
            grade = get_orm_session().query(Grade).get(student.grade_id) if student.grade_id else None
            
            formatted_students.append(format_student_details(student, user, school, class_obj, grade))
        
        return jsonify(formatted_students), 200

    except SQLAlchemyError as e:
        logging.error(f"Database error while fetching students by school and class: {str(e)}")
        return jsonify({"message": "Internal server error while querying data", "details": str(e)}), 500
    except Exception as e:
        logging.error(f"Unexpected error in get_students_by_school_and_class route: {str(e)}", exc_info=True)
        return jsonify({"message": "An unexpected error occurred", "details": str(e)}), 500

@bp.route('/<string:student_id>/class', methods=['GET'])
@jwt_required()
@role_required("admin", "diretor", "coordenador", "professor", "aluno")
def get_student_class(student_id):
    try:
        logging.info(f"Fetching class for user_id: {student_id}")

        # Buscar o aluno com suas relações
        student = get_orm_session().query(
            Student,
            User,
            School,
            Class,
            Grade
        ).join(
            User, Student.user_id == User.id
        ).join(
            School, Student.school_id == School.id
        ).join(
            Class, Student.class_id == Class.id
        ).join(
            Grade, Student.grade_id == Grade.id
        ).filter(
            Student.user_id == student_id
        ).first()

        if not student:
            logging.warning(f"Student not found with user_id: {student_id}")
            return jsonify({"message": "Student not found"}), 404

        student_obj, user, school, class_obj, grade = student

        # Verificar permissões do usuário
        current_user = get_current_user_from_token()
        if not current_user:
            return jsonify({"message": "User not found"}), 404

        if current_user['role'] == "professor":
            # Professor só pode ver alunos da escola onde está alocado
            if current_user.get('school_id') != student_obj.school_id:
                logging.warning(f"Professor {current_user.get('id')} tried to access student from different school")
                return jsonify({"message": "You don't have permission to view this student's class"}), 403
        elif current_user['role'] in ["diretor", "coordenador"]:
            # Diretor e coordenador só podem ver alunos de escolas da sua cidade
            city_id = current_user.get('tenant_id') or current_user.get('city_id')
            if not city_id:
                return jsonify({"message": "City ID not available"}), 400
            if school.city_id != city_id:
                logging.warning(f"User {current_user.get('id')} tried to access student from different city")
                return jsonify({"message": "You don't have permission to view this student's class"}), 403

        # Buscar avaliações aplicadas à classe usando o relacionamento class_tests
        class_tests = get_orm_session().query(ClassTest).filter_by(class_id=class_obj.id).all()
        applied_test_ids = [ct.test_id for ct in class_tests]

        # Formatar resposta com detalhes da classe
        class_details = {
            "id": class_obj.id,
            "name": class_obj.name,
            "school_id": class_obj.school_id,
            "grade_id": str(class_obj.grade_id) if class_obj.grade_id else None,
            "school": {
                "id": school.id,
                "name": school.name,
                "domain": school.domain,
                "address": school.address,
                "city_id": school.city_id
            },
            "grade": {
                "id": grade.id,
                "name": grade.name
            } if grade else None,
            "student": {
                "id": student_obj.id,
                "name": student_obj.name,
                "registration": student_obj.registration,
                "birth_date": student_obj.birth_date.isoformat() if student_obj.birth_date else None,
                "user_id": student_obj.user_id
            },
            "applied_tests": {
                "total": len(applied_test_ids),
                "test_ids": applied_test_ids
            }
        }

        return jsonify(class_details), 200

    except SQLAlchemyError as e:
        logging.error(f"Database error while fetching student class: {str(e)}")
        return jsonify({"message": "Internal server error while querying data", "details": str(e)}), 500
    except Exception as e:
        logging.error(f"Unexpected error in get_student_class route: {str(e)}", exc_info=True)
        return jsonify({"message": "An unexpected error occurred", "details": str(e)}), 500

@bp.route('/me', methods=['GET'])
@jwt_required()
@role_required("aluno")
def get_current_student():
    """
    Retorna os dados completos do aluno atual baseado no token JWT
    Inclui: nome completo, escola, estado, município, turma, série e professores
    """
    try:
        from app.utils.tenant_middleware import get_current_tenant_context, set_search_path, city_id_to_schema_name

        user = get_current_user_from_token()
        if not user:
            return jsonify({"error": "Usuário não encontrado"}), 404

        # Garantir schema do tenant: Student está em city_xxx, não em public
        tenant_ctx = get_current_tenant_context()
        if not tenant_ctx or not getattr(tenant_ctx, "has_tenant_context", False) or tenant_ctx.schema == "public":
            user_obj = get_orm_session().query(User).get(user["id"])
            if user_obj and getattr(user_obj, "city_id", None):
                schema = city_id_to_schema_name(str(user_obj.city_id))
                set_search_path(schema)
            else:
                return jsonify({
                    "error": "Contexto do município não disponível. Acesse pelo subdomínio da cidade (ex: cidade.afirmeplay.com.br) ou verifique se o usuário está vinculado a um município."
                }), 400

        # Buscar aluno pelo user_id com todas as relações
        student = get_orm_session().query(
            Student,
            User,
            School,
            Class,
            Grade
        ).join(
            User, Student.user_id == User.id
        ).outerjoin(
            School, Student.school_id == School.id
        ).outerjoin(
            Class, Student.class_id == Class.id
        ).outerjoin(
            Grade, Student.grade_id == Grade.id
        ).filter(
            Student.user_id == user['id']
        ).first()

        if not student:
            return jsonify({"error": "Dados do aluno não encontrados"}), 404

        student_obj, user_obj, school, class_obj, grade = student

        # Buscar município e estado através da escola
        city = None
        if school and school.city_id:
            city = get_orm_session().query(City).get(school.city_id)

        # Buscar professores vinculados à turma do aluno
        teachers_data = []
        if class_obj:
            teacher_classes = get_orm_session().query(TeacherClass).filter_by(class_id=class_obj.id).all()
            
            for tc in teacher_classes:
                teacher = get_orm_session().query(Teacher).get(tc.teacher_id)
                if teacher:
                    teacher_user = get_orm_session().query(User).get(teacher.user_id)
                    teachers_data.append({
                        "id": teacher.id,
                        "name": teacher.name,
                        "email": teacher_user.email if teacher_user else None,
                        "registration": teacher.registration,
                        "user_id": teacher.user_id
                    })

        # Montar resposta completa (mesmo formato da rota /students/<student_id>)
        response_data = {
            "id": student_obj.id,
            "name": student_obj.name or user_obj.name,
            "full_name": user_obj.name,  # Nome completo do usuário
            "email": user_obj.email,
            "registration": student_obj.registration or user_obj.registration,
            "birth_date": student_obj.birth_date.isoformat() if student_obj.birth_date else None,
            "address": user_obj.address,  # Endereço do aluno
            "created_at": student_obj.created_at.isoformat() if student_obj.created_at else None,
            "school": {
                "id": school.id,
                "name": school.name,
                "address": school.address,
                "domain": school.domain
            } if school else None,
            "city": {
                "id": city.id,
                "name": city.name,
                "state": city.state
            } if city else None,
            "municipio": city.name if city else None,  # Nome do município
            "estado": city.state if city else None,  # Estado
            "class": {
                "id": class_obj.id,
                "name": class_obj.name,
                "school_id": class_obj.school_id
            } if class_obj else None,
            "turma": {
                "id": class_obj.id,
                "name": class_obj.name,
                "school_id": class_obj.school_id
            } if class_obj else None,  # Turma atrelada
            "grade": {
                "id": str(grade.id),
                "name": grade.name
            } if grade else None,
            "serie": {
                "id": str(grade.id),
                "name": grade.name
            } if grade else None,  # Série
            "teachers": teachers_data,  # Professores vinculados à turma
            "professores": teachers_data  # Alias para professores
        }

        return jsonify(response_data), 200

    except SQLAlchemyError as e:
        logging.error(f"Erro no banco de dados ao obter dados do aluno atual: {str(e)}")
        return jsonify({"message": "Erro interno do servidor ao consultar dados", "details": str(e)}), 500
    except Exception as e:
        logging.error(f"Erro ao obter dados do aluno atual: {str(e)}", exc_info=True)
        return jsonify({"error": "Erro ao obter dados do aluno", "details": str(e)}), 500

@bp.route('/password-report', methods=['GET'])
@jwt_required()
@role_required("admin", "professor", "coordenador", "diretor", "tecadm")
def get_password_report():
    """
    Retorna relatório de alunos com suas senhas em formato Excel
    
    Query Parameters:
        city_id: Filtrar por cidade (opcional)
        school_id: Filtrar por escola (opcional)
        class_id: Filtrar por turma (opcional)
        grade_id: Filtrar por série (opcional)
        date_from: Data inicial (opcional, formato: YYYY-MM-DD)
        date_to: Data final (opcional, formato: YYYY-MM-DD)
    
    Returns:
        Arquivo Excel (.xlsx) com relatório de senhas
    """
    try:
        # 1. Obter usuário atual
        user = get_current_user_from_token()
        if not user:
            return jsonify({"error": "Usuário não encontrado"}), 401
        
        # 2. Construir query base
        query = get_orm_session().query(
            StudentPasswordLog,
            School,
            City,
            Class,
            Grade
        ).outerjoin(
            School, StudentPasswordLog.school_id == School.id
        ).outerjoin(
            City, StudentPasswordLog.city_id == City.id
        ).outerjoin(
            Class, StudentPasswordLog.class_id == Class.id
        ).outerjoin(
            Grade, StudentPasswordLog.grade_id == Grade.id
        )
        
        # 3. Aplicar filtros automáticos baseados no role
        if user['role'] == "admin":
            # Admin pode ver todos os alunos
            pass
        elif user['role'] == "tecadm":
            # Tecadm vê apenas alunos do seu município
            city_id = user.get('tenant_id') or user.get('city_id')
            if not city_id:
                return jsonify({"error": "ID da cidade não disponível para este usuário"}), 400
            query = query.filter(StudentPasswordLog.city_id == city_id)
        elif user['role'] in ["diretor", "coordenador"]:
            # Diretor e coordenador veem apenas alunos da sua escola
            manager = get_orm_session().query(Manager).filter_by(user_id=user['id']).first()
            if not manager:
                return jsonify({"error": "Diretor/Coordenador não encontrado na tabela manager"}), 404
            if not manager.school_id:
                return jsonify({"error": "Diretor/Coordenador não está vinculado a nenhuma escola"}), 400
            query = query.filter(StudentPasswordLog.school_id == manager.school_id)
        elif user['role'] == "professor":
            # Professor vê apenas alunos das escolas onde está vinculado
            teacher = get_orm_session().query(Teacher).filter_by(user_id=user['id']).first()
            if not teacher:
                return jsonify({"error": "Professor não encontrado na tabela teacher"}), 404
            
            # Buscar escolas onde o professor está vinculado
            school_teachers = get_orm_session().query(SchoolTeacher).filter_by(teacher_id=teacher.id).all()
            school_ids = [st.school_id for st in school_teachers]
            
            if not school_ids:
                return jsonify({"error": "Professor não está vinculado a nenhuma escola"}), 400
            
            # Aplicar filtro por escolas
            query = query.filter(StudentPasswordLog.school_id.in_(school_ids))
            
            # Buscar turmas onde o professor está vinculado
            teacher_classes = get_orm_session().query(TeacherClass).filter_by(teacher_id=teacher.id).all()
            class_ids = [tc.class_id for tc in teacher_classes]
            
            # Se professor está vinculado a turmas específicas, filtrar por elas
            # Se não houver turmas vinculadas, mostrar todas as turmas das escolas vinculadas
            if class_ids:
                query = query.filter(StudentPasswordLog.class_id.in_(class_ids))
        
        # 4. Aplicar filtros opcionais (query parameters)
        city_id_param = request.args.get('city_id')
        school_id_param = request.args.get('school_id')
        class_id_param = request.args.get('class_id')
        grade_id_param = request.args.get('grade_id')
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        
        if city_id_param:
            query = query.filter(StudentPasswordLog.city_id == city_id_param)
        if school_id_param:
            query = query.filter(StudentPasswordLog.school_id == school_id_param)
        if class_id_param:
            query = query.filter(StudentPasswordLog.class_id == class_id_param)
        if grade_id_param:
            query = query.filter(StudentPasswordLog.grade_id == grade_id_param)
        if date_from:
            try:
                date_from_obj = datetime.strptime(date_from, "%Y-%m-%d").date()
                query = query.filter(StudentPasswordLog.created_at >= date_from_obj)
            except ValueError:
                return jsonify({"error": "Formato de data inválido para date_from. Use YYYY-MM-DD"}), 400
        if date_to:
            try:
                date_to_obj = datetime.strptime(date_to, "%Y-%m-%d").date()
                # Adicionar 23:59:59 para incluir todo o dia
                date_to_datetime = datetime.combine(date_to_obj, datetime.max.time())
                query = query.filter(StudentPasswordLog.created_at <= date_to_datetime)
            except ValueError:
                return jsonify({"error": "Formato de data inválido para date_to. Use YYYY-MM-DD"}), 400
        
        # 5. Ordenar por data de criação (mais recentes primeiro)
        query = query.order_by(StudentPasswordLog.created_at.desc())
        
        # 6. Executar query
        results = query.all()
        
        if not results:
            return jsonify({"error": "Nenhum registro encontrado com os filtros aplicados"}), 404
        
        # 7. Gerar Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório de Senhas"
        
        # Criar cabeçalho
        headers = ["Nome do Aluno", "Email", "Senha", "Matrícula", "Escola", "Cidade", "Série", "Turma", "Data de Criação"]
        ws.append(headers)
        
        # Estilizar cabeçalho
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Adicionar dados
        for log, school, city, class_obj, grade in results:
            created_at_str = ""
            if log.created_at:
                if isinstance(log.created_at, datetime):
                    created_at_str = log.created_at.strftime("%d/%m/%Y %H:%M:%S")
                else:
                    created_at_str = str(log.created_at)
            
            ws.append([
                log.student_name or "",
                log.email or "",
                log.password,  # Senha em texto plano
                log.registration or "",
                school.name if school else "",
                city.name if city else "",
                grade.name if grade else "",
                class_obj.name if class_obj else "",
                created_at_str
            ])
        
        # Ajustar largura das colunas
        column_widths = {
            'A': 30,  # Nome do Aluno
            'B': 30,  # Email
            'C': 30,  # Senha
            'D': 15,  # Matrícula
            'E': 30,  # Escola
            'F': 25,  # Cidade
            'G': 15,  # Série
            'H': 20,  # Turma
            'I': 20   # Data de Criação
        }
        
        for column, width in column_widths.items():
            ws.column_dimensions[column].width = width
        
        # 8. Salvar em memória
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # 9. Gerar nome do arquivo com timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"relatorio_senhas_alunos_{timestamp}.xlsx"
        
        # 10. Retornar arquivo Excel
        response = Response(
            buffer.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        logging.info(f"Relatório de senhas gerado com sucesso. Total de registros: {len(results)}")
        
        return response
        
    except Exception as e:
        logging.error(f"Erro ao gerar relatório Excel de senhas: {str(e)}", exc_info=True)
        return jsonify({"error": "Erro ao gerar relatório", "details": str(e)}), 500


@bp.route('/password-report/pdf', methods=['GET'])
@jwt_required()
@role_required("admin", "professor", "coordenador", "diretor", "tecadm")
def get_password_report_pdf():
    """
    Retorna relatório de alunos com e-mails e senhas em formato PDF, agrupado por turma.
    Inclui capa com nome da escola e tabelas por turma (Nome, E-mail, Senha, Matrícula).

    Query Parameters:
        school_id: Escola (obrigatório para admin/tecadm; diretor/coordenador usam a escola vinculada)
        grade_id: Filtrar por série (opcional) – todas as turmas dessa série
        class_id: Filtrar por turma (opcional) – apenas uma turma

    Returns:
        Arquivo PDF com capa e tabelas por turma.
    """
    try:
        user = get_current_user_from_token()
        if not user:
            return jsonify({"error": "Usuário não encontrado"}), 401

        # Construir query base (mesma do Excel)
        query = get_orm_session().query(
            StudentPasswordLog,
            School,
            City,
            Class,
            Grade
        ).outerjoin(
            School, StudentPasswordLog.school_id == School.id
        ).outerjoin(
            City, StudentPasswordLog.city_id == City.id
        ).outerjoin(
            Class, StudentPasswordLog.class_id == Class.id
        ).outerjoin(
            Grade, StudentPasswordLog.grade_id == Grade.id
        )

        # Filtros por role
        school_id_param = request.args.get('school_id')
        if user['role'] == "admin":
            if not school_id_param:
                return jsonify({"error": "Para gerar o PDF é necessário informar school_id"}), 400
            query = query.filter(StudentPasswordLog.school_id == school_id_param)
        elif user['role'] == "tecadm":
            city_id = user.get('tenant_id') or user.get('city_id')
            if not city_id:
                return jsonify({"error": "ID da cidade não disponível para este usuário"}), 400
            query = query.filter(StudentPasswordLog.city_id == city_id)
            if school_id_param:
                query = query.filter(StudentPasswordLog.school_id == school_id_param)
        elif user['role'] in ["diretor", "coordenador"]:
            manager = get_orm_session().query(Manager).filter_by(user_id=user['id']).first()
            if not manager or not manager.school_id:
                return jsonify({"error": "Diretor/Coordenador não está vinculado a nenhuma escola"}), 400
            query = query.filter(StudentPasswordLog.school_id == manager.school_id)
            school_id_param = str(manager.school_id)
        elif user['role'] == "professor":
            teacher = get_orm_session().query(Teacher).filter_by(user_id=user['id']).first()
            if not teacher:
                return jsonify({"error": "Professor não encontrado"}), 404
            school_teachers = get_orm_session().query(SchoolTeacher).filter_by(teacher_id=teacher.id).all()
            school_ids = [st.school_id for st in school_teachers]
            if not school_ids:
                return jsonify({"error": "Professor não está vinculado a nenhuma escola"}), 400
            query = query.filter(StudentPasswordLog.school_id.in_(school_ids))
            if len(school_ids) > 1 and not school_id_param:
                return jsonify({"error": "Professor vinculado a mais de uma escola. Informe school_id para gerar o PDF."}), 400
            if not school_id_param:
                school_id_param = school_ids[0]
            else:
                query = query.filter(StudentPasswordLog.school_id == school_id_param)

        class_id_param = request.args.get('class_id')
        grade_id_param = request.args.get('grade_id')
        if class_id_param:
            query = query.filter(StudentPasswordLog.class_id == class_id_param)
        if grade_id_param:
            query = query.filter(StudentPasswordLog.grade_id == grade_id_param)

        # Ordenar por série, turma e nome para agrupamento
        query = query.order_by(
            Grade.name.asc().nullslast(),
            Class.name.asc().nullslast(),
            StudentPasswordLog.student_name.asc()
        )
        results = query.all()

        # Agrupar por turma (class_id + class_name + grade_name)
        turmas_map = {}
        for log, school, city, class_obj, grade in results:
            class_id_key = str(log.class_id) if log.class_id else "_sem_turma_"
            class_name = class_obj.name if class_obj else "—"
            grade_name = grade.name if grade else "—"
            key = (class_id_key, class_name, grade_name)
            if key not in turmas_map:
                turmas_map[key] = {"class_name": class_name, "grade_name": grade_name, "alunos": []}
            turmas_map[key]["alunos"].append({
                "student_name": log.student_name,
                "email": log.email,
                "password": log.password,
                "registration": log.registration,
            })

        turmas = list(turmas_map.values())

        # Metadados para a capa
        escola_nome = None
        municipio_nome = None
        serie_label = None
        if results:
            _, school, city, _, grade = results[0]
            escola_nome = school.name if school else None
            municipio_nome = city.name if city else None
            if grade_id_param and grade:
                serie_label = grade.name
            elif grade:
                serie_label = grade.name
        if not escola_nome and school_id_param:
            school_obj = get_orm_session().query(School).get(school_id_param)
            escola_nome = school_obj.name if school_obj else None
            if school_obj and school_obj.city_id:
                city_obj = get_orm_session().query(City).get(school_obj.city_id)
                municipio_nome = city_obj.name if city_obj else None

        total_alunos = sum(len(t["alunos"]) for t in turmas)
        data_geracao = datetime.now().strftime("%d/%m/%Y %H:%M")

        metadados = {
            "titulo": "Relatório de Acesso dos Alunos",
            "escola": escola_nome,
            "municipio": municipio_nome,
            "data_geracao": data_geracao,
            "serie_label": serie_label if grade_id_param else None,
            "total_alunos": total_alunos,
            "show_logo": False,
        }

        templates_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'templates')
        app_dir = os.path.dirname(os.path.dirname(__file__))
        temp_dir = tempfile.mkdtemp()
        try:
            logo_src = _get_logo_path()
            if logo_src:
                logo_dest = os.path.join(temp_dir, 'afirme_logo.png')
                shutil.copy2(logo_src, logo_dest)
                metadados["show_logo"] = True
        except Exception as e:
            logging.warning(f"Logo não copiada para temp: {e}")

        env = Environment(loader=FileSystemLoader(templates_dir), autoescape=select_autoescape(['html', 'xml']))
        template = env.get_template('student_passwords_report.html')
        html_content = template.render(metadados=metadados, turmas=turmas)

        temp_html_path = os.path.join(temp_dir, 'report.html')
        with open(temp_html_path, 'w', encoding='utf-8') as f:
            f.write(html_content)

        try:
            pdf_content = HTML(
                filename=temp_html_path,
                base_url=temp_dir
            ).write_pdf()
        finally:
            try:
                shutil.rmtree(temp_dir, ignore_errors=True)
            except Exception:
                pass

        safe_escola = (escola_nome or "escola").replace(" ", "_").replace("/", "_")[:40]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"relatorio_acesso_alunos_{safe_escola}_{timestamp}.pdf"

        response = make_response(pdf_content)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
        logging.info(f"Relatório PDF de senhas gerado. Escola: {escola_nome}, Alunos: {total_alunos}")
        return response

    except Exception as e:
        logging.error(f"Erro ao gerar relatório PDF de senhas: {str(e)}", exc_info=True)
        return jsonify({"error": "Erro ao gerar relatório PDF", "details": str(e)}), 500
