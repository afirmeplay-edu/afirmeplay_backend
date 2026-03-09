from flask import Blueprint, jsonify, request
from app.decorators.role_required import get_current_tenant_id
from app.models.user import User, RoleEnum
from app.decorators.role_required import role_required, get_current_user_from_token
from flask_jwt_extended import jwt_required
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy import func, text
import logging
from werkzeug.security import generate_password_hash, check_password_hash
from app import db
from app.models.student import Student
from app.models.school import School
from app.models.studentClass import Class
from app.models.grades import Grade
from app.models.teacher import Teacher
from app.models.manager import Manager
from app.models.city import City
from app.models.studentPasswordLog import StudentPasswordLog
from app.models.user_settings import UserSettings
from sqlalchemy.orm import joinedload
from app.utils.email_service import EmailService
from app.utils.tenant_middleware import city_id_to_schema_name, get_current_tenant_context
from datetime import datetime, timedelta, date
from flask import current_app
import csv
import io
import uuid
from werkzeug.utils import secure_filename
import os
from openpyxl import load_workbook
from xlrd import open_workbook
import re

bp = Blueprint('users', __name__, url_prefix='/users')


def normalizar_nome_para_busca(nome):
    """
    Normaliza um nome para busca, removendo espaços extras e convertendo para minúsculas.
    Remove espaços no início/fim e normaliza espaços múltiplos para um único espaço.
    """
    if not nome:
        return ""
    # Converter para string, remover espaços no início/fim e normalizar espaços múltiplos
    nome_normalizado = re.sub(r'\s+', ' ', str(nome).strip())
    return nome_normalizado.lower()


def gerar_iniciais_email(nome_completo):
    """
    Gera a parte local do email com a primeira letra de cada nome.
    Ex: "Artur Alexandre Calderon" -> "aac"
    """
    if not nome_completo:
        return ""
    partes = re.sub(r'\s+', ' ', str(nome_completo).strip()).split()
    return "".join(p[0].lower() for p in partes if p)


def gerar_primeiro_nome(nome_completo):
    """
    Retorna o primeiro nome em minúsculas.
    Ex: "Artur Alexandre Calderon" -> "artur"
    """
    if not nome_completo:
        return ""
    partes = re.sub(r'\s+', ' ', str(nome_completo).strip()).split()
    return partes[0].lower() if partes else ""


def obter_email_disponivel(base_local, dominio="@afirmeplay.com.br", usados_nesse_batch=None):
    """
    Retorna um email disponível: base_local + dominio.
    Se base_local@dominio já existir (no banco ou no batch), usa base_local1, base_local2, etc.
    """
    usados_nesse_batch = usados_nesse_batch or set()
    candidato = f"{base_local}{dominio}"
    if candidato not in usados_nesse_batch and not User.query.filter(func.lower(User.email) == candidato.lower()).first():
        return candidato
    suffix = 1
    while True:
        candidato = f"{base_local}{suffix}{dominio}"
        if candidato not in usados_nesse_batch and not User.query.filter(func.lower(User.email) == candidato.lower()).first():
            return candidato
        suffix += 1

def format_student_details(student):
    user_details = None
    school_details = None
    if student.school:
        school_details = {
            "id": student.school.id,
            "name": student.school.name,
            "domain": student.school.domain,
            "address": student.school.address,
            "city_id": student.school.city_id,
            "created_at": student.school.created_at.isoformat() if student.school.created_at else None
        }

    class_details = None
    if student.class_:
        class_details = {
            "id": student.class_.id,
            "name": student.class_.name,
            "school_id": student.class_.school_id,
            "grade_id": str(student.class_.grade_id) if student.class_.grade_id else None
        }

    grade_details = None
    if student.grade:
        grade_details = {
            "id": student.grade.id,
            "name": student.grade.name
        }

    return {
        "id": student.id,
        "name": student.name,
        "registration": student.registration,
        "birth_date": student.birth_date.isoformat() if student.birth_date else None,
        "class_id": student.class_id,
        "grade_id": student.grade_id,
        "school_id": student.school_id,
        "created_at": student.created_at.isoformat() if student.created_at else None,
        "school": school_details,
        "class": class_details,
        "grade": grade_details
    }


def format_user_settings(settings_instance):
    if not settings_instance:
        return {
            "theme": None,
            "fontFamily": None,
            "fontSize": None,
            "sidebar_theme_id": None,
            "frame_id": None,
            "stamp_id": None,
        }
    return {
        "theme": settings_instance.theme,
        "fontFamily": settings_instance.font_family,
        "fontSize": settings_instance.font_size,
        "sidebar_theme_id": settings_instance.sidebar_theme_id,
        "frame_id": settings_instance.frame_id,
        "stamp_id": settings_instance.stamp_id,
    }

@bp.route('/list', methods=['GET'])
@jwt_required()
@role_required("admin", "tecadm", "diretor", "coordenador", "professor")
def list_users():
    try:
        current_user = get_current_user_from_token()
        if not current_user:
            return jsonify({"erro": "Usuário não encontrado"}), 404

        # Base query
        query = User.query

        # Filtragem baseada no papel do usuário
        if current_user['role'] == "admin":
            # Admin vê todos os usuários do sistema
            pass
        elif current_user['role'] == "tecadm":
            # Tecadm vê apenas usuários do seu município
            city_id = current_user.get('tenant_id') or current_user.get('city_id')
            if not city_id:
                return jsonify({"erro": "ID da cidade não disponível"}), 400
            query = query.filter_by(city_id=city_id)
        elif current_user['role'] in ["diretor", "coordenador"]:
            # Diretor e coordenador vêem apenas usuários da sua escola
            from app.models.manager import Manager
            from app.models.school import School
            
            # Buscar o manager vinculado ao usuário atual
            manager = Manager.query.filter_by(user_id=current_user['id']).first()
            if not manager or not manager.school_id:
                return jsonify({"erro": "Usuário não está vinculado a nenhuma escola"}), 400
            
            # Buscar a escola do manager
            school = School.query.get(manager.school_id)
            if not school:
                return jsonify({"erro": "Escola não encontrada"}), 400
            
            # Filtrar usuários por city_id da escola
            query = query.filter_by(city_id=school.city_id)
        elif current_user['role'] == "professor":
            # Professor vê apenas usuários do seu município
            city_id = current_user.get('city_id')
            if not city_id:
                return jsonify({"erro": "ID da cidade não disponível"}), 400
            query = query.filter_by(city_id=city_id)

        users = query.all()
        
        return jsonify({
            "users": [{
                "id": user.id,
                "name": user.name,
                "email": user.email,
                "registration": user.registration,
                "role": user.role.value,
                "city_id": user.city_id,
                "created_at": user.created_at.isoformat() if user.created_at else None,
                "updated_at": user.updated_at.isoformat() if user.updated_at else None
            } for user in users]
        }), 200

    except SQLAlchemyError as e:
        logging.error(f"Erro no banco de dados ao listar usuários: {e}")
        return jsonify({"erro": "Erro ao consultar usuários", "detalhes": str(e)}), 500
    except Exception as e:
        logging.error(f"Erro inesperado ao listar usuários: {e}", exc_info=True)
        return jsonify({"erro": "Erro interno do servidor"}), 500

@bp.route('', methods=['POST'])
@jwt_required()
@role_required("admin", "professor", "coordenador", "diretor", "tecadm")
def create_user():
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "No data provided"}), 400

        required_fields = ["name", "email", "password"]
        for field in required_fields:
            if field not in data:
                return jsonify({"error": f"Missing required field: {field}"}), 400

        # Obter usuário atual
        current_user = get_current_user_from_token()
        if not current_user:
            return jsonify({"error": "User not found"}), 404

        # Check if email already exists
        existing_user = User.query.filter_by(email=data["email"]).first()
        if existing_user:
            return jsonify({
                "message": "User already exists",
                "user": {
                    "id": existing_user.id,
                    "name": existing_user.name,
                    "email": existing_user.email,
                    "registration": existing_user.registration,
                    "role": existing_user.role.value
                }
            }), 200

        # Check if registration already exists
        if data.get("registration") and User.query.filter_by(registration=data["registration"]).first():
            return jsonify({"error": "Registration number already exists"}), 400

        # Determinar city_id baseado na role do usuário atual
        city_id = None
        if current_user['role'] == "admin":
            # Admin pode escolher qualquer city_id
            city_id = data.get("city_id")
            if not city_id:
                return jsonify({"error": "city_id is required for admin creating users"}), 400
        else:
            # TecAdmin, Diretor, Coordenador, Professor usam seu próprio city_id
            city_id = current_user.get("city_id")
            if not city_id:
                return jsonify({"error": "User does not have a city_id assigned"}), 400

        # Para alunos, não definir city_id diretamente (será herdado da escola)
        if data.get("role") == "aluno":
            city_id = None

        # Para tecadm, city_id é obrigatório
        if data.get("role") == "tecadm":
            city_id = data.get("city_id")
            if not city_id:
                return jsonify({"error": "city_id is required for tecadm"}), 400

        # Create user
        novo_usuario = User(
            name=data["name"],
            email=data["email"],
            password_hash=generate_password_hash(data["password"]),
            registration=data.get("registration"),
            role=RoleEnum(data.get("role")),
            city_id=city_id  # Adicionando city_id
        )
        db.session.add(novo_usuario)
        db.session.commit()

        return jsonify({
            "message": "User created successfully!",
            "user": {
                "id": novo_usuario.id,
                "name": novo_usuario.name,
                "email": novo_usuario.email,
                "registration": novo_usuario.registration,
                "role": novo_usuario.role.value,
                "city_id": novo_usuario.city_id
            }
        }), 201

    except SQLAlchemyError as e:
        db.session.rollback()
        logging.error(f"Database error: {str(e)}")
        return jsonify({"error": "Database error occurred", "details": str(e)}), 500
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error creating user: {str(e)}", exc_info=True)
        return jsonify({"error": "Error creating user", "details": str(e)}), 500

def serialize_user(user):
    role_val = None
    if getattr(user, 'role', None) is not None:
        role_val = getattr(user.role, 'value', user.role)
        if callable(role_val):
            role_val = None
    return {
        "id": user.id,
        "name": user.name,
        "email": user.email,
        "registration": user.registration,
        "role": role_val,
        "city_id": user.city_id,
        "birth_date": user.birth_date.isoformat() if user.birth_date else None,
        "nationality": user.nationality,
        "phone": user.phone,
        "gender": user.gender,
        "traits": user.traits,
        "avatar_config": user.avatar_config,
        "address": user.address,
        "created_at": user.created_at.isoformat() if user.created_at else None,
        "updated_at": user.updated_at.isoformat() if user.updated_at else None
    }


def _user_needs_onboarding(user):
    """Considera que o usuário precisa de onboarding se não tiver onboarding_completed em avatar_config."""
    if not user:
        return True
    ac = user.avatar_config
    if not ac or not isinstance(ac, dict):
        return True
    return ac.get("onboarding_completed") is not True


@bp.route('/me/onboarding-status', methods=['GET'])
@jwt_required()
def get_onboarding_status():
    """
    Retorna se o usuário atual precisa preencher o modal de primeiro acesso (onboarding).
    Usado pelo frontend para exibir o modal e pré-preencher os campos.
    Sem migration: usa avatar_config.onboarding_completed (JSON existente).
    """
    try:
        current_user = get_current_user_from_token()
        if not current_user:
            return jsonify({"erro": "Usuário não encontrado"}), 401

        user = User.query.get(current_user["id"])
        if not user:
            return jsonify({"erro": "Usuário não encontrado"}), 404

        needs_onboarding = _user_needs_onboarding(user)
        profile = {
            "name": user.name,
            "birth_date": user.birth_date.isoformat() if user.birth_date else None,
            "phone": user.phone,
            "gender": user.gender,
            "nationality": user.nationality,
            "address": user.address,
            "traits": user.traits,
            "avatar_config": user.avatar_config,
        }
        return jsonify({
            "needs_onboarding": needs_onboarding,
            "profile": profile,
        }), 200
    except Exception as e:
        logging.error(f"Erro ao obter status de onboarding: {str(e)}", exc_info=True)
        return jsonify({"erro": "Erro ao obter status de onboarding", "detalhes": str(e)}), 500


@bp.route('/me/onboarding', methods=['POST'])
@jwt_required()
def submit_onboarding():
    """
    Envio do formulário do modal de primeiro acesso.
    Atualiza: name, birth_date, phone, gender, nationality, address, traits, avatar_config
    e marca onboarding_completed em avatar_config (sem migration).
    """
    try:
        current_user = get_current_user_from_token()
        if not current_user:
            return jsonify({"erro": "Usuário não encontrado"}), 401

        user = User.query.get(current_user["id"])
        if not user:
            return jsonify({"erro": "Usuário não encontrado"}), 404

        data = request.get_json() or {}

        # name (opcional no onboarding)
        name = data.get("name")
        if name is not None and isinstance(name, str) and name.strip():
            user.name = name.strip()

        # birth_date
        birth_date_value = data.get("birth_date")
        if birth_date_value:
            try:
                if isinstance(birth_date_value, str):
                    user.birth_date = datetime.strptime(birth_date_value, "%Y-%m-%d").date()
                elif isinstance(birth_date_value, date):
                    user.birth_date = birth_date_value
                else:
                    return jsonify({"erro": "birth_date deve estar no formato YYYY-MM-DD"}), 400
            except ValueError:
                return jsonify({"erro": "birth_date inválido"}), 400

        # nationality
        nationality = data.get("nationality")
        if nationality is not None:
            user.nationality = nationality.strip() if isinstance(nationality, str) and nationality.strip() else None

        # phone
        phone = data.get("phone")
        if phone is not None:
            if phone:
                phone_digits = re.sub(r"\D", "", str(phone))
                user.phone = phone_digits if phone_digits and phone_digits.isdigit() else None
            else:
                user.phone = None

        # gender
        gender = data.get("gender")
        allowed_genders = {"masculino", "feminino", "outro", "prefiro_nao_informar"}
        if gender is not None:
            if gender:
                g = gender.strip().lower()
                if g not in allowed_genders:
                    return jsonify({"erro": "gender inválido"}), 400
                user.gender = g
            else:
                user.gender = None

        # traits (características - lista de strings)
        traits = data.get("traits")
        if traits is not None:
            if traits in ("", [], None):
                user.traits = None
            elif isinstance(traits, list) and all(isinstance(t, str) for t in traits):
                user.traits = traits
            else:
                return jsonify({"erro": "traits deve ser uma lista de strings ou null"}), 400

        # address
        address = data.get("address")
        if address is not None:
            user.address = address.strip() if isinstance(address, str) and address.strip() else None

        # avatar_config: theme (modo escuro), font, icon + marcar onboarding_completed
        avatar_config = dict(user.avatar_config) if user.avatar_config and isinstance(user.avatar_config, dict) else {}
        if "theme" in data:
            v = data["theme"]
            if v in ("light", "dark", "system"):
                avatar_config["theme"] = v
        if "font" in data and data["font"]:
            avatar_config["font"] = str(data["font"]).strip()
        if "icon" in data:
            avatar_config["icon"] = data["icon"]  # string ou número do ícone
        avatar_config["onboarding_completed"] = True
        user.avatar_config = avatar_config

        db.session.commit()

        return jsonify({
            "message": "Configuração concluída. Você já pode acessar o sistema.",
            "user": serialize_user(user),
        }), 200
    except SQLAlchemyError as e:
        db.session.rollback()
        logging.error(f"Erro ao salvar onboarding: {str(e)}")
        return jsonify({"erro": "Erro ao salvar configuração", "detalhes": str(e)}), 500
    except Exception as e:
        db.session.rollback()
        logging.error(f"Erro inesperado no onboarding: {str(e)}", exc_info=True)
        return jsonify({"erro": "Erro ao salvar configuração"}), 500


@bp.route('/<string:user_id>', methods=['GET'])
@jwt_required()
@role_required("admin", "tecadm", "diretor", "coordenador", "professor", "aluno")
def get_user_by_id(user_id):
    try:
        logging.info(f"Fetching user with ID: {user_id}")

        user = User.query.get(user_id)

        if not user:
            logging.warning(f"User not found with ID: {user_id}")
            return jsonify({"message": "User not found"}), 404

        user_data = serialize_user(user)

        if getattr(user, 'role', None) == RoleEnum.ALUNO:
            student = None
            try:
                from app.utils.tenant_middleware import ensure_tenant_schema_for_user
                if ensure_tenant_schema_for_user(user.id):
                    student = Student.query.options(
                        joinedload(Student.school),
                        joinedload(Student.class_),
                        joinedload(Student.grade)
                    ).filter_by(user_id=user.id).first()
            except Exception as student_err:
                logging.warning(f"Erro ao buscar student_details para user {user_id}: {student_err}")
            if student:
                try:
                    student_details = format_student_details(student)
                    user_data['student_details'] = student_details
                except Exception as fmt_err:
                    logging.warning(f"Erro ao formatar student_details para user {user_id}: {fmt_err}", exc_info=True)
                    user_data['student_details'] = None

        return jsonify(user_data), 200

    except SQLAlchemyError as e:
        logging.error(f"Database error while fetching user by ID: {str(e)}")
        return jsonify({"message": "Internal server error while querying data", "details": str(e)}), 500
    except Exception as e:
        logging.error(f"Unexpected error in get_user_by_id route: {str(e)}", exc_info=True)
        return jsonify({"message": "An unexpected error occurred", "details": str(e)}), 500

@bp.route('/forgot-password', methods=['POST'])
def forgot_password():
    """Solicita redefinição de senha"""
    try:
        data = request.get_json()
        if not data or 'email' not in data:
            return jsonify({"erro": "Email é obrigatório"}), 400
        
        email = data['email'].strip().lower()
        
        # Buscar usuário pelo email
        user = User.query.filter_by(email=email).first()
        
        if not user:
            # Por segurança, não revelar se o email existe ou não
            return jsonify({
                "mensagem": "Se o email estiver cadastrado em nossa base, você receberá um link para redefinir sua senha."
            }), 200
        
        # Gerar token de reset
        email_service = EmailService()
        reset_token = email_service.generate_reset_token()
        
        # Calcular expiração (1 hora)
        expiry_time = datetime.utcnow() + timedelta(seconds=current_app.config.get('PASSWORD_RESET_TOKEN_EXPIRY', 3600))
        
        # Salvar token no banco
        user.reset_token = reset_token
        user.reset_token_expires = expiry_time
        db.session.commit()
        
        # Enviar email
        email_sent = email_service.send_password_reset_email(
            user_email=user.email,
            user_name=user.name,
            reset_token=reset_token
        )
        
        if email_sent:
            return jsonify({
                "mensagem": "Se o email estiver cadastrado em nossa base, você receberá um link para redefinir sua senha."
            }), 200
        else:
            # Se falhou ao enviar email, limpar token
            user.reset_token = None
            user.reset_token_expires = None
            db.session.commit()
            
            return jsonify({
                "erro": "Erro ao enviar email. Tente novamente mais tarde."
            }), 500
            
    except SQLAlchemyError as e:
        db.session.rollback()
        logging.error(f"Erro no banco de dados ao solicitar reset de senha: {e}")
        return jsonify({"erro": "Erro interno do servidor"}), 500
    except Exception as e:
        logging.error(f"Erro inesperado ao solicitar reset de senha: {e}", exc_info=True)
        return jsonify({"erro": "Erro interno do servidor"}), 500

@bp.route('/reset-password', methods=['POST'])
def reset_password():
    """Redefine a senha usando token"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({"erro": "Dados não fornecidos"}), 400
        
        required_fields = ["token", "new_password"]
        for field in required_fields:
            if field not in data:
                return jsonify({"erro": f"Campo obrigatório: {field}"}), 400
        
        token = data['token']
        new_password = data['new_password']
        
        # Validar senha
        if len(new_password) < 6:
            return jsonify({"erro": "A senha deve ter pelo menos 6 caracteres"}), 400
        
        # Buscar usuário pelo token
        user = User.query.filter_by(reset_token=token).first()
        
        if not user:
            return jsonify({"erro": "Token inválido"}), 400
        
        # Verificar se o token expirou
        if user.reset_token_expires and user.reset_token_expires < datetime.utcnow():
            # Limpar token expirado
            user.reset_token = None
            user.reset_token_expires = None
            db.session.commit()
            return jsonify({"erro": "Token expirado. Solicite um novo link de redefinição."}), 400
        
        # Atualizar senha
        user.password_hash = generate_password_hash(new_password)
        user.reset_token = None
        user.reset_token_expires = None
        user.updated_at = datetime.utcnow()
        db.session.commit()
        
        # Enviar email de confirmação
        email_service = EmailService()
        email_service.send_password_changed_email(user.email, user.name)
        
        return jsonify({
            "mensagem": "Senha redefinida com sucesso!"
        }), 200
        
    except SQLAlchemyError as e:
        db.session.rollback()
        logging.error(f"Erro no banco de dados ao redefinir senha: {e}")
        return jsonify({"erro": "Erro interno do servidor"}), 500
    except Exception as e:
        logging.error(f"Erro inesperado ao redefinir senha: {e}", exc_info=True)
        return jsonify({"erro": "Erro interno do servidor"}), 500

@bp.route('/change-password', methods=['POST'])
@jwt_required()
def change_password():
    """Altera a senha do usuário logado"""
    try:
        current_user = get_current_user_from_token()
        if not current_user:
            return jsonify({"erro": "Usuário não encontrado"}), 404
        
        data = request.get_json()
        if not data:
            return jsonify({"erro": "Dados não fornecidos"}), 400
        
        required_fields = ["current_password", "new_password"]
        for field in required_fields:
            if field not in data:
                return jsonify({"erro": f"Campo obrigatório: {field}"}), 400
        
        current_password = data['current_password']
        new_password = data['new_password']
        
        # Validar senha
        if len(new_password) < 6:
            return jsonify({"erro": "A senha deve ter pelo menos 6 caracteres"}), 400
        
        # Buscar usuário
        user = User.query.get(current_user['id'])
        if not user:
            return jsonify({"erro": "Usuário não encontrado"}), 404
        
        # Verificar senha atual
        if not check_password_hash(user.password_hash, current_password):
            return jsonify({"erro": "Senha atual incorreta"}), 400
        
        # Verificar se a nova senha é diferente da atual
        if check_password_hash(user.password_hash, new_password):
            return jsonify({"erro": "A nova senha deve ser diferente da senha atual"}), 400
        
        # Atualizar senha
        user.password_hash = generate_password_hash(new_password)
        user.updated_at = datetime.utcnow()
        db.session.commit()
        
        # Enviar email de confirmação
        email_service = EmailService()
        email_service.send_password_changed_email(user.email, user.name)
        
        return jsonify({
            "mensagem": "Senha alterada com sucesso!"
        }), 200
        
    except SQLAlchemyError as e:
        db.session.rollback()
        logging.error(f"Erro no banco de dados ao alterar senha: {e}")
        return jsonify({"erro": "Erro interno do servidor"}), 500
    except Exception as e:
        logging.error(f"Erro inesperado ao alterar senha: {e}", exc_info=True)
        return jsonify({"erro": "Erro interno do servidor"}), 500

@bp.route('/check-email', methods=['POST'])
@jwt_required()
@role_required("admin", "tecadm", "diretor", "coordenador", "professor")
def check_email_availability():
    """Verifica se um email está disponível e sugere uma alternativa caso já esteja em uso"""
    try:
        data = request.get_json()
        if not data or 'email' not in data:
            return jsonify({"erro": "O campo 'email' é obrigatório"}), 400

        email = data['email'].strip().lower()

        if '@' not in email:
            return jsonify({"erro": "Email inválido"}), 400

        prefix, domain = email.split('@', 1)

        if not prefix:
            return jsonify({"erro": "Email inválido"}), 400

        existing = User.query.filter(func.lower(User.email) == email).first()

        if not existing:
            return jsonify({
                "disponivel": True,
                "email": email
            }), 200

        # Encontrar o próximo email disponível (ex: aac2, aac3, ...)
        suffix = 2
        while True:
            candidate = f"{prefix}{suffix}@{domain}"
            if not User.query.filter(func.lower(User.email) == candidate).first():
                return jsonify({
                    "disponivel": False,
                    "email": email,
                    "email_sugerido": candidate
                }), 200
            suffix += 1

    except SQLAlchemyError as e:
        logging.error(f"Erro no banco de dados ao verificar email: {e}")
        return jsonify({"erro": "Erro interno do servidor"}), 500
    except Exception as e:
        logging.error(f"Erro inesperado ao verificar email: {e}", exc_info=True)
        return jsonify({"erro": "Erro interno do servidor"}), 500


@bp.route('/validate-reset-token', methods=['POST'])
def validate_reset_token():
    """Valida se um token de reset é válido"""
    try:
        data = request.get_json()
        if not data or 'token' not in data:
            return jsonify({"erro": "Token é obrigatório"}), 400
        
        token = data['token']
        
        # Buscar usuário pelo token
        user = User.query.filter_by(reset_token=token).first()
        
        if not user:
            return jsonify({"valido": False, "mensagem": "Token inválido"}), 200
        
        # Verificar se o token expirou
        if user.reset_token_expires and user.reset_token_expires < datetime.utcnow():
            # Limpar token expirado
            user.reset_token = None
            user.reset_token_expires = None
            db.session.commit()
            return jsonify({"valido": False, "mensagem": "Token expirado"}), 200
        
        return jsonify({
            "valido": True,
            "mensagem": "Token válido",
            "email": user.email
        }), 200
        
    except SQLAlchemyError as e:
        logging.error(f"Erro no banco de dados ao validar token: {e}")
        return jsonify({"erro": "Erro interno do servidor"}), 500
    except Exception as e:
        logging.error(f"Erro inesperado ao validar token: {e}", exc_info=True)
        return jsonify({"erro": "Erro interno do servidor"}), 500 

@bp.route('/<string:user_id>', methods=['PUT'])
@jwt_required()
@role_required("admin", "tecadm", "diretor", "coordenador", "professor", "aluno")
def update_user(user_id):
    try:
        current_user = get_current_user_from_token()
        if not current_user:
            return jsonify({"erro": "Usuário não encontrado"}), 404

        if current_user['id'] != user_id and current_user['role'] not in ["admin", "tecadm"]:
            return jsonify({"erro": "Sem permissão para atualizar este usuário"}), 403

        user = User.query.get(user_id)
        if not user:
            return jsonify({"erro": "Usuário não encontrado"}), 404

        data = request.get_json() or {}

        # Atualização de campos básicos
        name = data.get('name')
        if name is not None:
            if isinstance(name, str) and name.strip():
                user.name = name.strip()
            else:
                return jsonify({"erro": "name deve ser uma string não vazia"}), 400

        email = data.get('email')
        if email is not None:
            if not isinstance(email, str) or not email.strip():
                return jsonify({"erro": "email deve ser uma string não vazia"}), 400
            email_clean = email.strip()
            if "@" not in email_clean:
                return jsonify({"erro": "email inválido"}), 400
            existing_email = User.query.filter(
                func.lower(User.email) == email_clean.lower(),
                User.id != user.id
            ).first()
            if existing_email:
                return jsonify({"erro": "Email já cadastrado"}), 400
            user.email = email_clean

        password = data.get('password')
        if password is not None:
            if not isinstance(password, str) or len(password) < 8:
                return jsonify({"erro": "password deve conter ao menos 8 caracteres"}), 400
            user.password_hash = generate_password_hash(password)

        registration = data.get('registration')
        if registration is not None:
            registration_clean = str(registration).strip() if registration else None
            if registration_clean:
                existing_registration = User.query.filter(
                    User.registration == registration_clean,
                    User.id != user.id
                ).first()
                if existing_registration:
                    return jsonify({"erro": "registration já cadastrada"}), 400
                user.registration = registration_clean
            else:
                user.registration = None

        city_id_value = data.get('city_id')
        if city_id_value is not None:
            if current_user['role'] not in ["admin", "tecadm"]:
                return jsonify({"erro": "Sem permissão para alterar city_id"}), 403

            if city_id_value:
                city = City.query.get(city_id_value)
                if not city:
                    return jsonify({"erro": "city_id inválido"}), 400
            else:
                city = None

            if current_user['role'] == "tecadm":
                current_city_id = current_user.get('tenant_id') or current_user.get('city_id')
                if city and city.id != current_city_id:
                    return jsonify({"erro": "Tecadm só pode atribuir usuários ao seu município"}), 403

            user.city_id = city.id if city else None

        role_value = data.get('role')
        if role_value is not None:
            if current_user['role'] not in ["admin", "tecadm"]:
                return jsonify({"erro": "Sem permissão para alterar role"}), 403
            try:
                new_role = RoleEnum(role_value)
            except ValueError:
                return jsonify({"erro": "role inválido"}), 400

            if current_user['role'] == "tecadm" and new_role == RoleEnum.ADMIN:
                return jsonify({"erro": "Tecadm não pode atribuir role admin"}), 403

            user.role = new_role

        birth_date_value = data.get('birth_date')
        if birth_date_value:
            try:
                if isinstance(birth_date_value, str):
                    user.birth_date = datetime.strptime(birth_date_value, '%Y-%m-%d').date()
                elif isinstance(birth_date_value, date):
                    user.birth_date = birth_date_value
                else:
                    return jsonify({"erro": "birth_date deve estar no formato YYYY-MM-DD"}), 400
            except ValueError:
                return jsonify({"erro": "birth_date deve estar no formato YYYY-MM-DD"}), 400
        elif birth_date_value is None:
            user.birth_date = None

        nationality = data.get('nationality')
        user.nationality = nationality.strip() if isinstance(nationality, str) and nationality.strip() else None

        phone = data.get('phone')
        if phone:
            phone_digits = re.sub(r'\D', '', str(phone))
            if phone_digits and not phone_digits.isdigit():
                return jsonify({"erro": "phone deve conter apenas dígitos"}), 400
            user.phone = phone_digits if phone_digits else None
        elif phone is None:
            user.phone = None

        gender = data.get('gender')
        allowed_genders = {"masculino", "feminino", "outro", "prefiro_nao_informar"}
        if gender:
            gender_value = gender.strip().lower()
            if gender_value not in allowed_genders:
                return jsonify({"erro": "gender inválido"}), 400
            user.gender = gender_value
        elif gender is None:
            user.gender = None

        traits = data.get('traits')
        if traits is not None:
            if traits == "" or traits == []:
                user.traits = None
            elif isinstance(traits, list) and all(isinstance(item, str) for item in traits):
                user.traits = traits
            else:
                return jsonify({"erro": "traits deve ser uma lista de strings ou null"}), 400

        avatar_config = data.get('avatar_config')
        if avatar_config is not None:
            if isinstance(avatar_config, dict):
                user.avatar_config = avatar_config
            else:
                return jsonify({"erro": "avatar_config deve ser um objeto"}), 400

        address = data.get('address')
        user.address = address.strip() if isinstance(address, str) and address.strip() else None

        db.session.commit()

        return jsonify({"user": serialize_user(user)}), 200

    except SQLAlchemyError as e:
        db.session.rollback()
        logging.error(f"Erro no banco de dados ao atualizar usuário {user_id}: {str(e)}")
        return jsonify({"erro": "Erro ao atualizar usuário", "detalhes": str(e)}), 500
    except Exception as e:
        db.session.rollback()
        logging.error(f"Erro inesperado ao atualizar usuário {user_id}: {str(e)}", exc_info=True)
        return jsonify({"erro": "Erro interno do servidor"}), 500


@bp.route('/<string:user_id>', methods=['DELETE'])
@jwt_required()
@role_required("admin", "tecadm")
def delete_user(user_id):
    """Deleta um usuário específico"""
    try:
        current_user = get_current_user_from_token()

        if not current_user:
            return jsonify({"erro": "Usuário não encontrado"}), 404

        if current_user['id'] == user_id:
            return jsonify({"erro": "Não é possível deletar o próprio usuário"}), 400

        user_to_delete = User.query.get(user_id)
        if not user_to_delete:
            return jsonify({"erro": "Usuário não encontrado"}), 404

        if user_to_delete.city_id:
            user_schema = city_id_to_schema_name(user_to_delete.city_id)
            search_path = f'"{user_schema}", public'
            db.session.execute(text(f"SET search_path TO {search_path}"))

        if current_user['role'] == "tecadm":
            current_city_id = current_user.get('tenant_id') or current_user.get('city_id')
            if user_to_delete.city_id != current_city_id:
                return jsonify({"erro": "Sem permissão para deletar usuário de outra cidade"}), 403

        deleted_relations = []

        teacher = Teacher.query.filter_by(user_id=user_id).first()
        teacher_id = teacher.id if teacher else None

        all_cities = City.query.all() if teacher_id else []

        if teacher_id:
            from app.models.teacherClass import TeacherClass
            total_deleted = 0
            for city in all_cities:
                city_schema = city_id_to_schema_name(city.id)
                db.session.execute(text(f'SET search_path TO "{city_schema}", public'))
                deleted = TeacherClass.query.filter_by(teacher_id=teacher_id).delete(synchronize_session=False)
                total_deleted += deleted
            if total_deleted > 0:
                deleted_relations.append(f"{total_deleted} vínculos com turmas")
            if user_to_delete.city_id:
                user_schema = city_id_to_schema_name(user_to_delete.city_id)
                db.session.execute(text(f'SET search_path TO "{user_schema}", public'))

        if teacher_id:
            from app.models.schoolTeacher import SchoolTeacher
            total_deleted = 0
            for city in all_cities:
                city_schema = city_id_to_schema_name(city.id)
                db.session.execute(text(f'SET search_path TO "{city_schema}", public'))
                deleted = SchoolTeacher.query.filter_by(teacher_id=teacher_id).delete(synchronize_session=False)
                total_deleted += deleted
            if total_deleted > 0:
                deleted_relations.append(f"{total_deleted} vínculos escolares")
            if user_to_delete.city_id:
                user_schema = city_id_to_schema_name(user_to_delete.city_id)
                db.session.execute(text(f'SET search_path TO "{user_schema}", public'))

        if teacher_id:
            db.session.flush()

        deleted_count = Student.query.filter_by(user_id=user_id).delete(synchronize_session=False)
        if deleted_count > 0:
            deleted_relations.append("estudante")

        # Remover logs de senha do aluno no schema da cidade (evita órfãos e duplicatas ao recadastrar)
        if user_to_delete.city_id:
            db.session.execute(text(f'SET search_path TO "{city_id_to_schema_name(user_to_delete.city_id)}", public'))
            deleted_logs = StudentPasswordLog.query.filter_by(user_id=user_id).delete(synchronize_session=False)
            if deleted_logs > 0:
                deleted_relations.append("log de senhas (student_password_log)")

        if teacher:
            total_deleted = 0
            for city in all_cities:
                city_schema = city_id_to_schema_name(city.id)
                db.session.execute(text(f'SET search_path TO "{city_schema}", public'))
                deleted = Teacher.query.filter_by(user_id=user_id).delete(synchronize_session=False)
                total_deleted += deleted
            if total_deleted > 0:
                deleted_relations.append(f"professor ({total_deleted} schemas)")
            if user_to_delete.city_id:
                user_schema = city_id_to_schema_name(user_to_delete.city_id)
                db.session.execute(text(f'SET search_path TO "{user_schema}", public'))

        deleted_count = Manager.query.filter_by(user_id=user_id).delete(synchronize_session=False)
        if deleted_count > 0:
            deleted_relations.append("manager")

        from app.models.userQuickLinks import UserQuickLinks
        deleted_count = UserQuickLinks.query.filter_by(user_id=user_id).delete(synchronize_session=False)
        if deleted_count > 0:
            deleted_relations.append("atalhos rápidos")

        db.session.flush()

        db.session.delete(user_to_delete)

        db.session.commit()


        if deleted_relations:
            relations_text = ", ".join(deleted_relations)
            message = f"Usuário deletado com sucesso. Registros relacionados também deletados: {relations_text}"
        else:
            message = "Usuário deletado com sucesso"

        return jsonify({
            "mensagem": message,
            "deleted_relations": deleted_relations
        }), 200

    except SQLAlchemyError as e:
        db.session.rollback()
        return jsonify({"erro": "Erro interno do servidor"}), 500
    except Exception as e:
        db.session.rollback()
        return jsonify({"erro": "Erro interno do servidor"}), 500

def _user_settings_in_public(fn):
    """Executa uma função com search_path em public para garantir leitura/escrita em public.user_settings (evita divergência multi-tenant)."""
    try:
        db.session.execute(text("SET search_path TO public"))
        return fn()
    finally:
        try:
            db.session.execute(text("SET search_path TO public"))
        except Exception:
            pass


@bp.route('/user-settings/<string:user_id>', methods=['POST'])
@jwt_required()
@role_required("admin", "tecadm", "diretor", "coordenador", "professor", "aluno")
def save_user_settings(user_id):
    try:
        current_user = get_current_user_from_token()
        if not current_user:
            return jsonify({"erro": "Usuário não encontrado"}), 404

        if current_user['id'] != user_id and current_user['role'] not in ["admin", "tecadm"]:
            return jsonify({"erro": "Sem permissão para atualizar preferências de outro usuário"}), 403

        user = User.query.get(user_id)
        if not user:
            return jsonify({"erro": "Usuário não encontrado"}), 404

        payload = request.get_json()
        if not payload or 'settings' not in payload:
            return jsonify({"erro": "Objeto settings é obrigatório"}), 400

        settings_data = payload.get('settings')
        if not isinstance(settings_data, dict):
            return jsonify({"erro": "Objeto settings deve ser um dicionário"}), 400

        def _parse_font_size(val):
            """Aceita int, '110' ou '110%' e retorna int (110) ou None."""
            if val is None:
                return None
            if isinstance(val, int):
                return val
            s = str(val).strip().rstrip('%')
            if not s:
                return None
            try:
                return int(s)
            except ValueError:
                return None

        if 'fontSize' in settings_data and settings_data['fontSize'] is not None:
            parsed = _parse_font_size(settings_data['fontSize'])
            if parsed is None:
                return jsonify({"erro": "fontSize deve ser um número inteiro ou percentual (ex: 110 ou 110%)"}), 400

        def _save():
            user_settings = UserSettings.query.filter_by(user_id=user_id).first()
            if not user_settings:
                user_settings = UserSettings(user_id=user_id)
                db.session.add(user_settings)

            # Atualização parcial: só altera os campos que vieram no body (evita apagar tema/fonte quando só envia sidebar_theme_id)
            if 'theme' in settings_data:
                user_settings.theme = settings_data.get('theme')
            if 'fontFamily' in settings_data:
                user_settings.font_family = settings_data.get('fontFamily')
            if 'fontSize' in settings_data:
                user_settings.font_size = _parse_font_size(settings_data['fontSize'])
            if 'sidebar_theme_id' in settings_data:
                user_settings.sidebar_theme_id = settings_data.get('sidebar_theme_id')
            if 'frame_id' in settings_data:
                user_settings.frame_id = settings_data.get('frame_id')
            if 'stamp_id' in settings_data:
                user_settings.stamp_id = settings_data.get('stamp_id')

            db.session.commit()
            return jsonify({"settings": format_user_settings(user_settings)}), 200

        return _user_settings_in_public(_save)

    except SQLAlchemyError as e:
        db.session.rollback()
        logging.error(f"Erro no banco de dados ao salvar preferências do usuário {user_id}: {str(e)}")
        return jsonify({"erro": "Erro ao salvar preferências", "detalhes": str(e)}), 500
    except Exception as e:
        db.session.rollback()
        logging.error(f"Erro inesperado ao salvar preferências do usuário {user_id}: {str(e)}", exc_info=True)
        return jsonify({"erro": "Erro interno do servidor"}), 500


@bp.route('/user-settings/<string:user_id>', methods=['GET'])
@jwt_required()
@role_required("admin", "tecadm", "diretor", "coordenador", "professor", "aluno")
def get_user_settings(user_id):
    try:
        current_user = get_current_user_from_token()
        if not current_user:
            return jsonify({"erro": "Usuário não encontrado"}), 404

        if current_user['id'] != user_id and current_user['role'] not in ["admin", "tecadm"]:
            return jsonify({"erro": "Sem permissão para visualizar preferências de outro usuário"}), 403

        user = User.query.get(user_id)
        if not user:
            return jsonify({"erro": "Usuário não encontrado"}), 404

        def _get():
            user_settings = UserSettings.query.filter_by(user_id=user_id).first()
            return jsonify({"settings": format_user_settings(user_settings)}), 200

        return _user_settings_in_public(_get)

    except SQLAlchemyError as e:
        logging.error(f"Erro no banco de dados ao buscar preferências do usuário {user_id}: {str(e)}")
        return jsonify({"erro": "Erro ao buscar preferências", "detalhes": str(e)}), 500
    except Exception as e:
        logging.error(f"Erro inesperado ao buscar preferências do usuário {user_id}: {str(e)}", exc_info=True)
        return jsonify({"erro": "Erro interno do servidor"}), 500


@bp.route('/bulk-upload-students', methods=['POST'])
@jwt_required()
@role_required("admin", "tecadm", "diretor", "coordenador")
def bulk_upload_students():
    """
    Rota para upload em massa de alunos via arquivo CSV ou Excel.
    O arquivo NÃO deve conter email nem senha: o sistema gera automaticamente
    email = primeira letra de cada nome + @afirmeplay.com.br (ex: aac@afirmeplay.com.br);
    senha = primeiro nome + @afirmeplay (ex: artur@afirmeplay).
    Em caso de iniciais repetidas, é acrescentado número (aac1, aac2, ...).
    Retorna nomes, emails e senhas criados.
    """
    try:
        # Verificar se há arquivo no request
        if 'file' not in request.files:
            return jsonify({"erro": "Nenhum arquivo enviado"}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"erro": "Nenhum arquivo selecionado"}), 400
        
        # Verificar extensão do arquivo
        allowed_extensions = {'csv', 'xlsx', 'xls'}
        file_extension = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
        
        if file_extension not in allowed_extensions:
            return jsonify({"erro": "Formato de arquivo não suportado. Use CSV, XLSX ou XLS"}), 400
        
        # Obter usuário atual para verificar permissões
        current_user = get_current_user_from_token()
        if not current_user:
            return jsonify({"erro": "Usuário não encontrado"}), 404
        
        # Exigir contexto de cidade: school/student existem só em city_xxx, não em public
        tenant_ctx = get_current_tenant_context()
        if not tenant_ctx or not getattr(tenant_ctx, 'has_tenant_context', False) or tenant_ctx.schema == 'public':
            return jsonify({
                "erro": "É necessário informar o município para esta operação. "
                        "Se for admin, envie o header X-City-ID (ou X-City-Slug). "
                        "Se for tecadm, diretor ou coordenador, faça login com usuário vinculado a uma cidade."
            }), 400
        
        # Ler o arquivo
        try:
            if file_extension == 'csv':
                # Para CSV, tentar diferentes encodings
                try:
                    content = file.read().decode('utf-8')
                except UnicodeDecodeError:
                    file.seek(0)
                    content = file.read().decode('latin-1')
                
                # Processar CSV
                csv_reader = csv.DictReader(io.StringIO(content))
                rows = list(csv_reader)
            else:
                # Para Excel
                if file_extension == 'xlsx':
                    workbook = load_workbook(file, data_only=True)
                    sheet = workbook.active
                else:  # xls
                    workbook = open_workbook(file_contents=file.read())
                    sheet = workbook.sheet_by_index(0)
                
                # Converter para lista de dicionários
                rows = []
                if file_extension == 'xlsx':
                    headers = [cell.value for cell in sheet[1]]
                    for row in sheet.iter_rows(min_row=2):
                        row_data = {}
                        for i, cell in enumerate(row):
                            if i < len(headers) and headers[i]:
                                row_data[headers[i]] = cell.value
                        if any(row_data.values()):  # Ignorar linhas vazias
                            rows.append(row_data)
                else:  # xls
                    headers = [sheet.cell_value(0, i) for i in range(sheet.ncols)]
                    for row_idx in range(1, sheet.nrows):
                        row_data = {}
                        for col_idx in range(sheet.ncols):
                            if col_idx < len(headers) and headers[col_idx]:
                                row_data[headers[col_idx]] = sheet.cell_value(row_idx, col_idx)
                        if any(row_data.values()):  # Ignorar linhas vazias
                            rows.append(row_data)
        except Exception as e:
            return jsonify({"erro": f"Erro ao ler arquivo: {str(e)}"}), 400
        
        # Verificar colunas obrigatórias (email e senha são gerados pelo sistema)
        required_columns = ['nome', 'data_nascimento', 'escola', 'endereco_escola', 'estado_escola', 'municipio_escola', 'serie', 'turma']
        if rows:
            missing_columns = [col for col in required_columns if col not in rows[0].keys()]
        else:
            missing_columns = required_columns
        
        if missing_columns:
            return jsonify({
                "erro": f"Colunas obrigatórias ausentes: {', '.join(missing_columns)}",
                "colunas_encontradas": list(rows[0].keys()) if rows else [],
                "colunas_obrigatorias": required_columns
            }), 400
        
        # Limpar dados (nome e dados da escola/turma obrigatórios; email e senha não vêm do arquivo)
        rows = [row for row in rows if all(str(row.get(col, '')).strip() for col in ['nome', 'escola', 'endereco_escola', 'estado_escola', 'municipio_escola', 'serie', 'turma'])]
        
        # Converter data de nascimento
        def parse_date(date_value):
            # Se o valor for None ou vazio, retornar None
            if not date_value:
                return None
            
            # Se o valor já for um objeto date, retornar diretamente
            if isinstance(date_value, date):
                return date_value
            
            # Se o valor for um objeto datetime.datetime, converter para date
            if isinstance(date_value, datetime):
                return date_value.date()
            
            # Se for string, fazer o parse como antes
            try:
                date_str = str(date_value).strip()
                if not date_str:
                    return None
                
                # Tentar diferentes formatos de data
                for fmt in ['%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y', '%d-%m-%y']:
                    try:
                        return datetime.strptime(date_str, fmt).date()
                    except:
                        continue
                
                # Se nenhum formato funcionar, tentar parse automático com YYYY-MM-DD
                return datetime.strptime(date_str, '%Y-%m-%d').date()
            except:
                return None
        
        # Processar cada linha para adicionar data parseada
        for row in rows:
            row['data_nascimento_parsed'] = parse_date(row.get('data_nascimento'))
        
        # Validar dados e processar
        results = {
            "total_linhas": len(rows),
            "sucessos": 0,
            "erros": [],
            "alunos_criados": []
        }
        
        # Verificar permissões baseadas no papel do usuário
        allowed_schools = []
        if current_user['role'] == "admin":
            # Admin pode criar alunos em qualquer escola
            allowed_schools = [school.id for school in School.query.all()]
        elif current_user['role'] == "tecadm":
            # Tecadm pode criar alunos em escolas da sua cidade
            city_id = current_user.get('tenant_id') or current_user.get('city_id')
            if city_id:
                allowed_schools = [school.id for school in School.query.filter_by(city_id=city_id).all()]
        elif current_user['role'] in ["diretor", "coordenador"]:
            # Diretor e coordenador podem criar alunos apenas na sua escola
            manager = Manager.query.filter_by(user_id=current_user['id']).first()
            if manager and manager.school_id:
                allowed_schools = [manager.school_id]
        
        if not allowed_schools:
            return jsonify({"erro": "Usuário não tem permissão para criar alunos em nenhuma escola"}), 403
        
        # Conjunto de emails já atribuídos neste batch (para evitar duplicatas)
        emails_usados_no_batch = set()
        
        # Processar cada linha
        for index, row in enumerate(rows):
            try:
                nome_completo = str(row.get('nome', '')).strip()
                if not nome_completo:
                    results["erros"].append({
                        "linha": index + 2,
                        "campo": "nome",
                        "valor": "",
                        "erro": "Nome é obrigatório"
                    })
                    continue
                
                # Gerar email: primeira letra de cada nome + @afirmeplay.com.br (com sufixo se duplicado)
                base_local = gerar_iniciais_email(nome_completo)
                if not base_local:
                    results["erros"].append({
                        "linha": index + 2,
                        "campo": "nome",
                        "valor": nome_completo,
                        "erro": "Não foi possível gerar iniciais para o email"
                    })
                    continue
                email = obter_email_disponivel(base_local, "@afirmeplay.com.br", emails_usados_no_batch)
                emails_usados_no_batch.add(email.lower())
                
                # Gerar senha: primeiro nome + @afirmeplay
                primeiro_nome = gerar_primeiro_nome(nome_completo)
                senha = f"{primeiro_nome}@afirmeplay" if primeiro_nome else "aluno@afirmeplay"
                
                # Validar matrícula se fornecida (opcional)
                matricula_raw = row.get('matricula', '')
                # Normalizar matrícula: tratar None, "None", "null", "", etc. como vazio
                if matricula_raw is None:
                    matricula = None
                else:
                    matricula = str(matricula_raw).strip()
                    # Se após strip for vazio ou valores como "None", "null", etc., tratar como None
                    if not matricula or matricula.lower() in ['none', 'null', 'nulo', '']:
                        matricula = None
                
                # Verificar se matrícula já existe nesta cidade (Student no schema atual)
                # Matrícula é única por cidade, não globalmente em User
                if matricula and Student.query.filter_by(registration=matricula).first():
                    results["erros"].append({
                        "linha": index + 2,
                        "campo": "matricula",
                        "valor": matricula,
                        "erro": "Matrícula já cadastrada nesta cidade"
                    })
                    continue
                
                # Buscar escola existente (busca normalizada - case-insensitive e espaços normalizados)
                escola_nome = str(row.get('escola', '')).strip()
                escola_nome_normalizado = normalizar_nome_para_busca(escola_nome)
                
                # Buscar todas as escolas e comparar com normalização
                todas_escolas = School.query.all()
                escola = None
                for escola_candidata in todas_escolas:
                    nome_escola_normalizado = normalizar_nome_para_busca(escola_candidata.name)
                    if nome_escola_normalizado == escola_nome_normalizado:
                        escola = escola_candidata
                        break
                
                if not escola:
                    results["erros"].append({
                        "linha": index + 2,
                        "campo": "escola",
                        "valor": escola_nome,
                        "erro": "Escola não encontrada"
                    })
                    continue
                
                # Verificar se usuário tem permissão para esta escola
                if escola.id not in allowed_schools:
                    results["erros"].append({
                        "linha": index + 2,
                        "campo": "escola",
                        "valor": escola_nome,
                        "erro": "Sem permissão para criar alunos nesta escola"
                    })
                    continue
                
                # Buscar série existente (busca normalizada - case-insensitive e espaços normalizados)
                serie_nome = str(row.get('serie', '')).strip()
                serie_nome_normalizado = normalizar_nome_para_busca(serie_nome)
                
                # Buscar todas as séries e comparar com normalização
                todas_series = Grade.query.all()
                serie = None
                for serie_candidata in todas_series:
                    nome_serie_normalizado = normalizar_nome_para_busca(serie_candidata.name)
                    if nome_serie_normalizado == serie_nome_normalizado:
                        serie = serie_candidata
                        break
                
                if not serie:
                    results["erros"].append({
                        "linha": index + 2,
                        "campo": "serie",
                        "valor": serie_nome,
                        "erro": "Série não encontrada"
                    })
                    continue
                
                # Buscar turma existente (busca normalizada - case-insensitive e espaços normalizados)
                turma_nome = str(row.get('turma', '')).strip()
                turma_nome_normalizado = normalizar_nome_para_busca(turma_nome)
                
                # Buscar turmas da escola e série, e comparar com normalização
                turmas_candidatas = Class.query.filter(
                    Class.school_id == escola.id,
                    Class.grade_id == serie.id
                ).all()
                
                turma = None
                for turma_candidata in turmas_candidatas:
                    nome_turma_normalizado = normalizar_nome_para_busca(turma_candidata.name)
                    if nome_turma_normalizado == turma_nome_normalizado:
                        turma = turma_candidata
                        break
                
                if not turma:
                    results["erros"].append({
                        "linha": index + 2,
                        "campo": "turma",
                        "valor": turma_nome,
                        "erro": f"Turma não encontrada na escola {escola_nome} e série {serie_nome}"
                    })
                    continue
                
                # Verificar se já existe aluno na mesma turma (por matrícula, se fornecida)
                if matricula:
                    existing_student_in_class = Student.query.filter_by(
                        class_id=turma.id,
                        registration=matricula
                    ).first()
                    if existing_student_in_class:
                        results["erros"].append({
                            "linha": index + 2,
                            "campo": "matricula",
                            "valor": matricula,
                            "erro": f"Aluno com esta matrícula já está cadastrado na turma {turma_nome}"
                        })
                        continue
                
                # Validar data de nascimento
                data_nascimento = row.get('data_nascimento_parsed')
                if not data_nascimento:
                    results["erros"].append({
                        "linha": index + 2,
                        "campo": "data_nascimento",
                        "valor": row.get('data_nascimento', ''),
                        "erro": "Data de nascimento inválida"
                    })
                    continue
                
                # Criar usuário: não gravar matrícula em User (única em public) para evitar
                # conflito entre cidades; matrícula fica só em Student (única por city schema)
                novo_usuario = User(
                    id=str(uuid.uuid4()),
                    name=str(row.get('nome', '')).strip(),
                    email=email,
                    password_hash=generate_password_hash(senha),
                    registration=None,  # matrícula apenas em Student (por cidade)
                    role=RoleEnum.ALUNO,
                    city_id=escola.city_id
                )
                db.session.add(novo_usuario)
                db.session.flush()  # Flush para obter o ID do usuário
                
                # Criar estudante
                novo_aluno = Student(
                    id=str(uuid.uuid4()),
                    name=str(row.get('nome', '')).strip(),
                    registration=matricula if matricula else None,
                    birth_date=data_nascimento,
                    profile_picture=str(row.get('foto_perfil', '')).strip() if row.get('foto_perfil') else None,
                    user_id=novo_usuario.id,
                    school_id=escola.id,
                    grade_id=serie.id,
                    class_id=turma.id
                )
                db.session.add(novo_aluno)
                db.session.flush()  # Flush para obter o ID do aluno
                
                # Salvar senha em texto plano na tabela de log (sem hash)
                password_log = StudentPasswordLog(
                    student_name=str(row.get('nome', '')).strip(),
                    email=email,
                    password=senha,  # Senha em texto plano (gerada pelo sistema)
                    registration=matricula if matricula else None,
                    user_id=novo_usuario.id,
                    student_id=novo_aluno.id,
                    class_id=turma.id,
                    grade_id=serie.id,
                    school_id=escola.id,
                    city_id=escola.city_id
                )
                db.session.add(password_log)
                
                # Commit para esta linha
                db.session.commit()
                
                results["sucessos"] += 1
                results["alunos_criados"].append({
                    "nome": novo_aluno.name,
                    "email": novo_usuario.email,
                    "senha": senha,
                    "matricula": novo_aluno.registration,
                    "escola": escola.name,
                    "serie": serie.name,
                    "turma": turma.name
                })
                
                logging.info(f"Aluno criado com sucesso: {novo_aluno.name} ({novo_usuario.email})")
                
            except Exception as e:
                db.session.rollback()
                results["erros"].append({
                    "linha": index + 2,
                    "campo": "geral",
                    "valor": "",
                    "erro": f"Erro inesperado: {str(e)}"
                })
                logging.error(f"Erro ao processar linha {index + 2}: {str(e)}")
                continue
        
        # Preparar resposta
        if results["sucessos"] > 0:
            message = f"Upload concluído! {results['sucessos']} alunos criados com sucesso."
            if results["erros"]:
                message += f" {len(results['erros'])} erros encontrados."
        else:
            message = "Nenhum aluno foi criado. Verifique os erros abaixo."
        
        return jsonify({
            "mensagem": message,
            "resumo": {
                "total_linhas": results["total_linhas"],
                "sucessos": results["sucessos"],
                "erros": len(results["erros"])
            },
            "alunos_criados": results["alunos_criados"],
            "erros": results["erros"]
        }), 200 if results["sucessos"] > 0 else 400
        
    except Exception as e:
        logging.error(f"Erro inesperado no upload em massa: {str(e)}", exc_info=True)
        return jsonify({"erro": "Erro interno do servidor", "detalhes": str(e)}), 500 